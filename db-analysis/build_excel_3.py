"""Build batch-3 30-query DBA tuning Excel: backtest analysis path + system support."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = Path(r"C:\Project\db-analysis\explain_results_3.txt")
OUT = Path(r"C:\Project\db-analysis\trading_db_tuning_local_30_batch3.xlsx")


def parse_explain_blocks(text):
    blocks = {}
    parts = re.split(r"^==== (Q\d+)\s+(.+)$", text, flags=re.MULTILINE)
    for i in range(1, len(parts), 3):
        qid, label, body = parts[i], parts[i + 1], parts[i + 2]
        m_exec = re.search(r"Execution Time:\s+([\d.]+)\s*ms", body)
        exec_ms = float(m_exec.group(1)) if m_exec else None
        plan_lines = []
        in_plan = False
        for ln in body.splitlines():
            if re.match(r"^-{5,}", ln):
                in_plan = True
                continue
            if in_plan:
                if ln.strip().startswith("Planning") or ln.strip().startswith("Execution Time"):
                    break
                if ln.strip():
                    plan_lines.append(ln)
        top_op = plan_lines[0].strip() if plan_lines else ""
        top_op_clean = re.sub(r"\s*\(cost=.+", "", top_op)
        flags = {
            "seq_scan_present": bool(re.search(r"->\s+(Parallel\s+)?Seq Scan on (?!compress_)", body))
                                or bool(re.match(r"^\s*Seq Scan on", plan_lines[0] if plan_lines else "")),
            "external_merge": bool(re.search(r"Sort Method:\s+external merge", body)),
            "bitmap_heap": bool(re.search(r"->\s+(Parallel\s+)?Bitmap Heap Scan", body)),
            "rows_removed_high": bool(re.search(r"Rows Removed by Filter:\s+(?:[1-9]\d{2,})", body)),
            "sort_present": bool(re.search(r"->\s+Sort", body)),
        }
        blocks[qid] = {"label": label.strip(), "exec_ms": exec_ms, "top_op": top_op_clean, "flags": flags}
    return blocks


text = RESULTS.read_text(encoding="utf-8-sig", errors="ignore")
parsed = parse_explain_blocks(text)
assert len(parsed) == 30, f"Expected 30, got {len(parsed)}"

QUERIES = [
    # ---- backtest_trade (243K rows) ----
    ("Q01", "backtestTradeRepository - findByBacktestRunIdAndStatus", "backtest_trade",
     "SELECT * FROM backtest_trade WHERE backtest_run_id=? AND status=?",
     None, "OK",
     "Bitmap Heap Scan via idx_backtest_trade_run_id + Filter status. 0.7ms. Status hampir selalu 'CLOSED' setelah run selesai, filter cheap. Partial index `(backtest_run_id) WHERE status='CLOSED'` bisa di-considered tapi cost saving marginal."),

    ("Q02", "backtestTradeRepository - findAllByBacktestRunId", "backtest_trade",
     "SELECT * FROM backtest_trade WHERE backtest_run_id=? ORDER BY entry_time ASC LIMIT 50000",
     "CREATE INDEX idx_backtest_trade_run_entry ON backtest_trade (backtest_run_id, entry_time);",
     "MEDIUM",
     "Bitmap Heap Scan + Sort. Sort masih in-memory karena rows per run kecil (~ratusan), tapi composite (backtest_run_id, entry_time) eliminate Sort step + dukung ORDER BY langsung dari index walk. Hot path /backtest/{id}/trades."),

    # ---- backtest_equity_point (788K rows) ----
    ("Q03", "backtestEquityPointRepository - findByBacktestRunIdOrderByEquityDateAsc", "backtest_equity_point",
     "SELECT * FROM backtest_equity_point WHERE backtest_run_id=? ORDER BY equity_date ASC LIMIT 50000",
     None, "OK",
     "Pakai UNIQUE INDEX uq_backtest_equity_point (backtest_run_id, equity_date) — composite sempurna untuk filter + sort. 3.3ms / 1235 rows."),

    # ---- backtest_trade_position (248K rows) ----
    ("Q04", "backtestTradePositionRepository - findAllByTradeId", "backtest_trade_position",
     "SELECT * FROM backtest_trade_position WHERE backtest_trade_id=? ORDER BY entry_time ASC LIMIT 100",
     None, "OK",
     "Pakai idx_backtest_trade_position_trade_id. Hit kecil (1-3 rows per trade), Sort tidak terlihat di plan. 1.6ms."),

    # ---- backtest_trade_event (0 rows, empty) ----
    ("Q05", "backtestTradeEventRepository - findAllByBacktestRunId", "backtest_trade_event",
     "SELECT * FROM backtest_trade_event WHERE backtest_run_id=? ORDER BY event_time ASC, created_time ASC LIMIT 200000",
     None, "OK",
     "Tabel kosong (0 rows). Index idx_backtest_trade_event_run_id ready. Tidak ada sort spill saat data masuk karena per-run rows bounded (200K cap)."),

    ("Q06", "backtestTradeEventRepository - findAllByTradeId", "backtest_trade_event",
     "SELECT * FROM backtest_trade_event WHERE trade_id=? ORDER BY event_time ASC, created_time ASC LIMIT 1000",
     "CREATE INDEX idx_backtest_trade_event_trade_id ON backtest_trade_event (trade_id, event_time);",
     "MEDIUM",
     "FEEDBACK DBA: tidak ada index pada `trade_id`. Saat ini tabel kosong, Seq Scan murah. Begitu data masuk (events ~2-4 per trade × ribuan trade per run), perlu composite (trade_id, event_time) — terutama karena query ORDER BY event_time."),

    ("Q07", "backtestTradeEventRepository - findAllByTradePositionId", "backtest_trade_event",
     "SELECT * FROM backtest_trade_event WHERE trade_position_id=? ORDER BY event_time ASC, created_time ASC LIMIT 1000",
     "CREATE INDEX idx_backtest_trade_event_position_id ON backtest_trade_event (trade_position_id, event_time);",
     "MEDIUM",
     "Sama dengan Q06: tidak ada index. Composite (trade_position_id, event_time) untuk read pattern position-detail UI."),

    # ---- strategy_daily_realized_curve (35 rows) ----
    ("Q08", "strategyDailyRealizedCurveRepository - findTopByAccountStrategyIdAndCurveDateBefore", "strategy_daily_realized_curve",
     "SELECT * FROM ... WHERE account_strategy_id=? AND curve_date<? ORDER BY curve_date DESC LIMIT 1",
     None, "OK",
     "Index Scan Backward pakai uq_strategy_daily_realized_curve (UNIQUE composite — likely (account_strategy_id, curve_date)). 0.18ms."),

    ("Q09", "strategyDailyRealizedCurveRepository - findByAccountStrategyIdAndCurveDate", "strategy_daily_realized_curve",
     "SELECT * FROM ... WHERE account_strategy_id=? AND curve_date=? LIMIT 1",
     None, "OK",
     "Exact match pakai uq_strategy_daily_realized_curve. 0.009ms. Sempurna."),

    ("Q10", "strategyDailyRealizedCurveRepository - findByCurveDate", "strategy_daily_realized_curve",
     "SELECT * FROM ... WHERE curve_date=? ORDER BY account_strategy_id ASC",
     None, "OK",
     "Seq Scan 35 rows. Filter leading dengan curve_date saja, leading column UNIQUE INDEX salah. Tidak ada index by curve_date saja — wajar pada tabel 35 rows. Begitu tumbuh: index `(curve_date)` recommended."),

    ("Q11", "strategyDailyRealizedCurveRepository - findLatestBeforeDateForStrategies (DISTINCT ON)", "strategy_daily_realized_curve",
     "SELECT DISTINCT ON (account_strategy_id) * FROM ... WHERE account_strategy_id IN(?) AND curve_date<? ORDER BY account_strategy_id, curve_date DESC",
     None, "OK",
     "DISTINCT ON + Seq Scan 35 rows. 0.023ms. Saat tumbuh, akan otomatis pakai uq_strategy_daily_realized_curve karena ORDER BY-nya match leading kolom."),

    ("Q12", "strategyDailyRealizedCurveRepository - findByAccountStrategyIdsAndCurveDate", "strategy_daily_realized_curve",
     "SELECT * FROM ... WHERE account_strategy_id IN(?) AND curve_date=?",
     None, "OK",
     "Seq Scan 35 rows. uq_strategy_daily_realized_curve akan ter-pakai saat data tumbuh."),

    ("Q13", "strategyDailyRealizedCurveRepository - findByAccountStrategyIdsAndCurveDateBetween", "strategy_daily_realized_curve",
     "SELECT * FROM ... WHERE account_strategy_id IN(?) AND curve_date BETWEEN ? AND ? ORDER BY account_strategy_id, curve_date ASC",
     None, "OK",
     "Seq Scan 35 rows. Korelasi pearson di CorrelationGuardService — bisa jadi hot path nanti. uq composite cover saat tumbuh."),

    ("Q14", "strategyDailyRealizedCurveRepository - findByAccountStrategyIdAndCurveDateBetween", "strategy_daily_realized_curve",
     "SELECT * FROM ... WHERE account_strategy_id=? AND curve_date BETWEEN ? AND ? ORDER BY curve_date ASC",
     None, "OK",
     "Index Scan on uq_strategy_daily_realized_curve. 0.01ms. PnlDeviationAlertService hot path."),

    # ---- walk_forward_run (5 rows) ----
    ("Q15", "walkForwardRunRepository - findFiltered", "walk_forward_run",
     "SELECT * FROM walk_forward_run WHERE (?strat IS NULL OR strategy_code=?) AND ... ORDER BY created_time DESC",
     None, "OK",
     "Seq Scan 5 rows. Saat ini optimal. Saat tumbuh > 1000 rows: existing idx_walk_forward_strategy + idx_walk_forward_verdict cover branch utama; ORDER BY created_time mungkin perlu index DESC, tapi tabel ini tidak akan tumbuh cepat (1 row per WF run)."),

    ("Q16", "walkForwardRunRepository - findRecentForStrategy", "walk_forward_run",
     "SELECT * FROM walk_forward_run WHERE strategy_code=? AND instrument=? AND interval_name=? ORDER BY created_time DESC LIMIT 20",
     None, "OK",
     "Pakai idx_walk_forward_strategy. 0.05ms. OK."),

    ("Q17", "walkForwardRunRepository - findByCreatedTimeAfter (derived)", "walk_forward_run",
     "SELECT * FROM walk_forward_run WHERE created_time > ? ORDER BY created_time DESC",
     None, "OK",
     "Seq Scan 5 rows. Phase 3 fold-trend scan. OK."),

    # ---- trade_positions (0 rows now) ----
    ("Q18", "tradePositionRepository - findByTradePositionId", "trade_positions",
     "SELECT * FROM trade_positions WHERE trade_position_id=?",
     None, "OK",
     "PK lookup. Index Scan on trade_positions_pkey. OK."),

    ("Q19", "tradePositionRepository - findAllByTradeId", "trade_positions",
     "SELECT * FROM trade_positions WHERE trade_id=? ORDER BY entry_time ASC",
     None, "OK",
     "Pakai idx_trade_positions_trade_id. Position per trade biasa 1-3 rows — Sort gratis. OK."),

    ("Q20", "tradePositionRepository - findAllByTradeIdAndStatus", "trade_positions",
     "SELECT * FROM trade_positions WHERE trade_id=? AND status=? ORDER BY entry_time ASC",
     None, "OK",
     "Index dari Q19 (idx_trade_positions_trade_id) cover; filter status post-lookup cheap karena cardinality per-trade rendah."),

    ("Q21", "tradePositionRepository - findAllByAssetAndStatus", "trade_positions",
     "SELECT * FROM trade_positions WHERE asset=? AND status=? ORDER BY entry_time ASC",
     "CREATE INDEX idx_trade_positions_asset_status ON trade_positions (asset, entry_time DESC) WHERE status IN ('OPEN','PARTIALLY_CLOSED');",
     "MEDIUM",
     "FEEDBACK DBA: tidak ada index pada `asset`. Saat ini Seq Scan (tabel kosong). Saat live trading aktif, query ini akan jadi hot — partial index pada status aktif (OPEN/PARTIALLY_CLOSED) hemat storage + dukung ORDER BY entry_time. Pattern sama dengan idx_trades_open_entry_time di batch 1."),

    ("Q22", "tradePositionRepository - findDailyClosedPositionAggregates", "trade_positions",
     "SELECT account_id, account_strategy_id, SUM(...), COUNT(*) FROM trade_positions WHERE status='CLOSED' AND exit_time>=? AND exit_time<? GROUP BY account_id, account_strategy_id",
     "CREATE INDEX idx_trade_positions_closed_exit ON trade_positions (exit_time) WHERE status='CLOSED';",
     "MEDIUM",
     "Daily aggregator. Saat ini Seq Scan tabel kosong. Pola sama dengan idx_trades_status_exit_time di batch 1. Partial index akan dipakai begitu trade_positions populated dari live trading."),

    # ---- idempotency_record (0 rows) ----
    ("Q23", "idempotencyRecordRepository - findUnexpired", "idempotency_record",
     "SELECT * FROM idempotency_record WHERE agent_name=? AND key=? AND expires_at>=?",
     None, "OK",
     "PK composite (agent_name, key) — direct lookup. 0.016ms. expires_at filter post-lookup cheap."),

    ("Q24", "idempotencyRecordRepository - deleteExpired (DELETE)", "idempotency_record",
     "DELETE FROM idempotency_record WHERE expires_at<?",
     None, "OK",
     "Existing idx_idempotency_record_expires_at akan dipakai saat data ada. Tabel kosong saat ini → Seq Scan. OK untuk TTL sweep — dijalankan periodic."),

    # ---- error_log (90 rows) ----
    ("Q25", "errorLogRepository - countOpen", "error_log",
     "SELECT COUNT(*) FROM error_log WHERE status IN ('NEW','INVESTIGATING') AND CASE severity WHEN ... THEN ... END >= :rank",
     None, "MEDIUM",
     "FEEDBACK DBA: CASE expression dalam WHERE clause **anti-index** — optimizer tidak bisa pakai composite indexing apa pun. Solusi sebenarnya bukan tambah index: (1) tambah kolom `severity_rank INT` (computed/stored), dengan partial index `(severity_rank) WHERE status IN ('NEW','INVESTIGATING')`. Atau (2) ubah caller selalu pass parameter sebagai severity string-array lalu pakai `severity = ANY(:severities)`. Saat ini Seq Scan 90 rows = 2.5ms, masih oke. Akan jadi bottleneck di error_log volume produksi."),

    ("Q26", "errorLogRepository - findOpenByFingerprint", "error_log",
     "SELECT * FROM error_log WHERE fingerprint=? AND status IN ('NEW','INVESTIGATING')",
     None, "OK",
     "Pakai uq_error_log_fingerprint_open (sepertinya UNIQUE partial WHERE status IN ('NEW','INVESTIGATING')). Sempurna untuk dedup. 0.012ms."),

    ("Q27", "errorLogRepository - bumpOccurrence (UPDATE)", "error_log",
     "UPDATE error_log SET occurrence_count=...+1, last_seen_at=? WHERE error_id=?",
     None, "OK",
     "Pakai error_log_pkey. 0.022ms. Hot path appender — sudah optimal."),

    ("Q28", "errorLogRepository - markNotified (UPDATE)", "error_log",
     "UPDATE error_log SET notified_at=?, notification_channels=? WHERE error_id=?",
     None, "OK",
     "Pakai error_log_pkey. 0.010ms."),

    ("Q29", "errorLogRepository - updateStatus (UPDATE)", "error_log",
     "UPDATE error_log SET status=?, resolved_at=?, resolved_by=? WHERE error_id=?",
     None, "OK",
     "Pakai error_log_pkey. 0.010ms."),

    # ---- alert_event (8849 rows) ----
    ("Q30", "alertEventRepository - findFiltered", "alert_event",
     "SELECT * FROM alert_event WHERE (?sev IS NULL OR severity=?) AND (?kind IS NULL OR kind=?) AND ... ORDER BY created_at DESC LIMIT 50",
     None, "MEDIUM",
     "FEEDBACK DBA: Index Scan on idx_alert_event_created (created_at DESC) tapi 16.6ms karena harus probe banyak rows untuk match kind filter. Filter `kind='ML_REGIME_GATE_BLOCKED'` selektif (4856 rows dari 8849), tapi index existing idx_alert_event_kind_created (kind, created_at DESC) **belum dipakai optimizer** untuk query ini karena pola `(?kind IS NULL OR kind=?)` block index access. Solusi: (A) split jadi 2 query method (with-kind vs without), atau (B) pakai Specification builder. Cara penanganan sama dengan Q11 batch 2 (findFiltered backtest_run)."),
]

assert len(QUERIES) == 30

# ----- Build Excel -----
wb = Workbook()
ws = wb.active
ws.title = "Tuning - trading_db (batch 3)"

HEADERS = [("No", 5), ("Repository - Method", 60), ("Target Table(s)", 26),
           ("Query (ringkas)", 70), ("Current Plan (top op)", 42),
           ("Exec Time (ms)", 12), ("Issues", 32), ("Recommended Index DDL", 90),
           ("Priority", 10), ("Remarks", 60), ("Feedback DBA", 80)]

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
priority_fills = {"HIGH": PatternFill("solid", fgColor="F8CBAD"),
                  "MEDIUM": PatternFill("solid", fgColor="FFE699"),
                  "LOW": PatternFill("solid", fgColor="E2EFDA"),
                  "OK": PatternFill("solid", fgColor="DDEBF7")}

for col_idx, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col_idx, value=label)
    c.fill, c.font = header_fill, header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "C2"

ISSUE_MESSAGES = {
    "external_merge": "Sort spill ke disk",
    "seq_scan_present": "Seq Scan terdeteksi",
    "bitmap_heap": "Bitmap Heap Scan",
    "rows_removed_high": "Banyak Rows Removed by Filter",
    "sort_present": "Sort step (mungkin bisa eliminate via composite)",
}

small_tables = {"strategy_daily_realized_curve", "walk_forward_run", "error_log",
                "backtest_trade_event", "trade_positions", "idempotency_record"}

for i, (qid, repo_method, target_table, sql, ddl, prio, dba_note) in enumerate(QUERIES, start=2):
    p = parsed[qid]
    issues = [msg for flag, msg in ISSUE_MESSAGES.items() if p["flags"].get(flag)]
    if target_table in small_tables and "Seq Scan terdeteksi" in issues:
        issues = [x for x in issues if x != "Seq Scan terdeteksi"]
        issues.append("Seq Scan (table kecil — wajar)")
    issue_text = "; ".join(issues) if issues else "—"
    row_vals = [int(qid[1:]), repo_method, target_table, sql, p["top_op"][:200],
                round(p["exec_ms"], 3) if p["exec_ms"] is not None else None,
                issue_text, ddl if ddl else "(tidak ada perubahan)", prio, repo_method, dba_note]
    for col_idx, v in enumerate(row_vals, start=1):
        c = ws.cell(row=i, column=col_idx, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        c.border = border
        c.font = Font(size=10)
    ws.cell(row=i, column=9).fill = priority_fills.get(prio, PatternFill())
    ws.cell(row=i, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=9).font = Font(bold=True, size=10)
    ws.row_dimensions[i].height = 95

# ----- Ringkasan -----
ws2 = wb.create_sheet("Ringkasan")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 90

prio_counts = {"HIGH": 0, "MEDIUM": 0, "OK": 0}
ddl_count = 0
for _, _, _, _, ddl, prio, _ in QUERIES:
    prio_counts[prio] = prio_counts.get(prio, 0) + 1
    if ddl:
        ddl_count += 1

summary_rows = [
    ("Database", "trading_db (PostgreSQL 16.13 + TimescaleDB 2.27)"),
    ("Batch", "3 — backtest analysis path + system support"),
    ("Analisa tanggal", "2026-05-24"),
    ("Total query dianalisa", "30"),
    ("Repositori dicakup", "BacktestTrade (2) · BacktestEquityPoint (1) · BacktestTradePosition (1) · BacktestTradeEvent (3) · StrategyDailyRealizedCurve (7) · WalkForwardRun (3) · TradePosition (5) · IdempotencyRecord (2) · ErrorLog (5) · AlertEvent (1)"),
    ("", ""),
    ("Klasifikasi prioritas", ""),
    ("HIGH (wajib tindakan)", f"{prio_counts['HIGH']} — (tidak ada — pola ini mostly read-after-event tables)"),
    ("MEDIUM (rekomendasi)", f"{prio_counts['MEDIUM']} — Q02 (BT trade sort), Q06+Q07 (BT event), Q21+Q22 (trade_positions), Q25 (errorLog CASE), Q30 (alert_event dynamic)"),
    ("OK (no change)", f"{prio_counts['OK']}"),
    ("Index DDL yang ditawarkan", f"{ddl_count}"),
    ("", ""),
    ("Catatan kunci", ""),
    ("Pola anti-index #1: CASE-in-WHERE", "Q25 errorLog.countOpen — `CASE severity WHEN ... END >= :rank`. Tidak bisa di-index. Real fix: kolom `severity_rank INT` computed, lalu partial index `(severity_rank) WHERE status IN (...)`."),
    ("Pola anti-index #2: dynamic OR-NULL", "Q30 alertEvent.findFiltered — pola `(?p IS NULL OR col=?)` block index access. Sama seperti Q11 batch 2. Fix: split method atau Spring Specification."),
    ("Empty tables — index ready", "trade_positions (0), backtest_trade_event (0), idempotency_record (0). Index existing menutup pattern utama; index baru direkomendasi siap saat live trading menghasilkan rows."),
    ("backtest big-3 sudah well-indexed", "backtest_trade (243K), backtest_equity_point (788K), backtest_trade_position (248K) semua punya idx_*_run_id atau composite UNIQUE. Q02 satu-satunya gain dari composite (run_id, entry_time)."),
    ("", ""),
    ("Lifecycle ke depan", ""),
    ("Setelah batch 3 di-review", "Apply 5 DDL ke local + prod. Plus diskusi schema-level: severity_rank di error_log (code change), Spring Specification untuk Q11/Q30 (code refactor)."),
    ("Batch 4 (kalau lanjut)", "Python services — research-orchestrator + ingest + inference + train. Beda surface (raw SQL embedded, bukan @Query). Akan butuh extract dari source files."),
    ("Total cumulative coverage", "90 queries dianalisa, 16 indeks applied/ditawarkan (7 batch 1 + 4 batch 2 + 5 batch 3). Coverage trading-engine @Query: ~57% (90 dari 157)."),
]
for r, (k, v) in enumerate(summary_rows, start=1):
    ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=11)
    ws2.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    if not k:
        ws2.cell(row=r, column=1).font = Font(size=11)
    if r <= 5 or k.startswith("HIGH") or k.startswith("MEDIUM") or k.startswith("OK") or k.startswith("Index"):
        ws2.row_dimensions[r].height = 22
    else:
        ws2.row_dimensions[r].height = 50

# ----- DDL sheet -----
ws3 = wb.create_sheet("Index DDL (siap-jalan)")
ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 120
ws3.cell(row=1, column=1, value="Q [Priority]").font = Font(bold=True)
ws3.cell(row=1, column=2, value="DDL — CREATE INDEX CONCURRENTLY (ONLINE, no lock)").font = Font(bold=True)

ddls = []
for qid, repo_method, _, _, ddl, prio, _ in QUERIES:
    if ddl:
        ddl_c = re.sub(r"CREATE (UNIQUE )?INDEX(?! CONCURRENTLY)",
                       lambda m: f"CREATE {m.group(1) or ''}INDEX CONCURRENTLY", ddl, count=1)
        ddls.append((qid, repo_method, prio, ddl_c))

for r, (qid, rm, prio, d) in enumerate(ddls, start=2):
    ws3.cell(row=r, column=1, value=f"{qid} [{prio}]").font = Font(size=10, bold=True)
    ws3.cell(row=r, column=2, value=f"-- {rm}\n{d}").alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    ws3.row_dimensions[r].height = 80

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"   Sheets: Tuning ({len(QUERIES)} rows) | Ringkasan | Index DDL ({len(ddls)} rows)")
print(f"   Priority: HIGH={prio_counts['HIGH']}, MEDIUM={prio_counts['MEDIUM']}, OK={prio_counts['OK']}")
