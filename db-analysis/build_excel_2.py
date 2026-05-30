"""Build batch-2 30-query DBA tuning Excel for trading_db (local Postgres)."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = Path(r"C:\Project\db-analysis\explain_results_2.txt")
OUT = Path(r"C:\Project\db-analysis\trading_db_tuning_local_30_batch2.xlsx")


def parse_explain_blocks(text: str) -> dict[str, dict]:
    blocks = {}
    parts = re.split(r"^==== (Q\d+)\s+(.+)$", text, flags=re.MULTILINE)
    for i in range(1, len(parts), 3):
        qid, label, body = parts[i], parts[i + 1], parts[i + 2]
        m_exec = re.search(r"Execution Time:\s+([\d.]+)\s*ms", body)
        exec_ms = float(m_exec.group(1)) if m_exec else None
        plan_lines = []
        in_plan = False
        for ln in body.splitlines():
            if re.match(r"^-{5,}", ln):
                in_plan = True
                continue
            if in_plan:
                if ln.strip().startswith("Planning") or ln.strip().startswith("Execution Time"):
                    break
                if ln.strip():
                    plan_lines.append(ln)
        top_op = plan_lines[0].strip() if plan_lines else ""
        top_op_clean = re.sub(r"\s*\(cost=.+", "", top_op)
        flags = {
            "seq_scan_present": bool(re.search(r"->\s+(Parallel\s+)?Seq Scan on (?!compress_)", body)) or bool(re.match(r"^\s*Seq Scan on", plan_lines[0] if plan_lines else "")),
            "external_merge": bool(re.search(r"Sort Method:\s+external merge", body)),
            "bitmap_heap": bool(re.search(r"->\s+(Parallel\s+)?Bitmap Heap Scan", body)),
            "rows_removed_high": bool(re.search(r"Rows Removed by Filter:\s+(?:[1-9]\d{2,})", body)),
        }
        blocks[qid] = {"label": label.strip(), "exec_ms": exec_ms, "top_op": top_op_clean, "flags": flags}
    return blocks


text = RESULTS.read_text(encoding="utf-8-sig", errors="ignore")
parsed = parse_explain_blocks(text)
assert len(parsed) == 30, f"Expected 30 queries, got {len(parsed)}"


QUERIES = [
    # ---- account_strategy (53 rows, 3 indexes) ----
    ("Q01", "accountStrategyRepository - findByIdForUpdate", "account_strategy",
     "SELECT a.* FROM account_strategy a WHERE a.account_strategy_id = ? FOR UPDATE",
     None, "OK",
     "PK index sudah ada (account_strategy_pkey). Saat ini Seq Scan karena tabel 53 rows — optimizer cost lebih murah dari index lookup. Akan otomatis pakai PK saat tumbuh."),

    ("Q02", "accountStrategyRepository - findByUniqueKey", "account_strategy",
     "SELECT acs.* FROM account_strategy WHERE account_id=? AND strategy_definition_id=? AND symbol=? AND interval_name=? AND is_deleted=false",
     "CREATE INDEX idx_account_strategy_uniq_lookup ON account_strategy (account_id, strategy_definition_id, symbol, interval_name) WHERE is_deleted=false;",
     "MEDIUM",
     "FEEDBACK DBA: JavaDoc menyebut 'uq_account_strategy' unique constraint, TAPI di pg_indexes tidak ditemukan. Tidak ada enforcement DB-level untuk uniqueness, dan tidak ada index buat lookup tuple-nya. Tabel masih 53 rows → Seq Scan saat ini OK. Index direkomendasi siap saat data tumbuh + memberikan uniqueness via UNIQUE INDEX (note: tambahkan UNIQUE jika invariant memang one-row-per-tuple)."),

    ("Q03", "accountStrategyRepository - findByEnabledTrueAndIntervalName", "account_strategy",
     "SELECT acs.* FROM account_strategy WHERE enabled=true AND is_deleted=false AND interval_name=? ORDER BY priority_order, created_time",
     None, "OK",
     "Live tick dispatch path. Seq Scan + Sort, 0.05ms pada 53 rows. Partial index existing idx_account_strategy_active (is_deleted=false) hanya pre-filter; cover-nya account_id. Untuk volume produksi tidak akan jadi bottleneck — strategi enabled jarang lebih dari ratusan."),

    ("Q04", "accountStrategyRepository - findAllEnabledStrategyRefs", "account_strategy",
     "SELECT account_id, account_strategy_id FROM account_strategy WHERE enabled=true AND is_deleted=false",
     None, "OK",
     "Metric exporter, 30s refresh. Projection 2-kolom mungil. Seq Scan optimal di tabel 53 rows."),

    ("Q05", "accountStrategyRepository - findAllStrategyCodeRefs", "account_strategy",
     "SELECT account_strategy_id, strategy_code FROM account_strategy WHERE is_deleted=false",
     None, "OK",
     "Metric exporter cache rebuild. 30s refresh. Projection 2-kolom. Seq Scan optimal."),

    ("Q06", "accountStrategyRepository - findAllPromoted", "account_strategy",
     "SELECT acs.* FROM account_strategy WHERE enabled=true AND simulated=false AND is_deleted=false",
     None, "OK",
     "PnL deviation alert. Filter 3-bool kombinasi → match sangat sedikit row. Seq Scan ok pada tabel 53 rows."),

    ("Q07", "accountStrategyRepository - findActivePreset", "account_strategy",
     "SELECT acs.* FROM account_strategy WHERE account_id=? AND strategy_definition_id=? AND symbol=? AND interval_name=? AND is_deleted=false AND enabled=true",
     None, "OK",
     "Index dari Q02 (idx_account_strategy_uniq_lookup) akan cover ini juga — leading kolom sama persis. Tidak butuh index terpisah."),

    ("Q08", "accountStrategyRepository - findAllVisibleToUser", "account_strategy + accounts",
     "SELECT acs.* FROM account_strategy JOIN accounts a ON a.account_id=acs.account_id WHERE acs.is_deleted=false AND (a.user_id=? OR acs.visibility='PUBLIC')",
     None, "OK",
     "Hash Join. 39 acs × 4 accounts = murah. Tabel masih kecil. Existing idx_account_strategy_public partial sudah cover branch visibility='PUBLIC'."),

    # ---- backtest_run (1058 rows, 6 indexes) ----
    ("Q09", "backtestRunRepository - countByUserIdAndStatusIn", "backtest_run",
     "SELECT COUNT(*) FROM backtest_run WHERE user_id=? AND status IN (?)",
     None, "OK",
     "Pakai idx_backtest_run_active_per_user (partial WHERE status IN PENDING/RUNNING). Sempurna untuk panggilan ini. Index Only Scan, 0.5ms."),

    ("Q10", "backtestRunRepository - findAllOrderByCreatedTimeDesc", "backtest_run",
     "SELECT * FROM backtest_run ORDER BY created_time DESC LIMIT ? OFFSET ?",
     "CREATE INDEX idx_backtest_run_created_time ON backtest_run (created_time DESC NULLS LAST);",
     "MEDIUM",
     "Seq Scan + Sort 1058 rows, 1.6ms. Untuk pagination admin yang sering dijalankan, index pada created_time DESC akan eliminate Sort + dukung LIMIT/OFFSET cepat. Index juga akan dipakai Q11+Q12 yang sort by created_time."),

    ("Q11", "backtestRunRepository - findFiltered", "backtest_run",
     "SELECT * FROM backtest_run WHERE (user_id=? OR triggered_by='RESEARCHER') AND ... ORDER BY created_time DESC LIMIT ? OFFSET ? (15 filter params)",
     None, "MEDIUM",
     "FEEDBACK DBA: Query 15-param dengan banyak `CAST(:p AS TEXT) IS NULL OR ...` branch. Sulit tuning generic — optimizer melihatnya sebagai filter complex. Saat ini 1.0ms karena tabel 1058 rows. Solusi non-index: pertimbangkan Specification/Criteria builder (Spring Data Specification) supaya bisa generate query bebas-NULL-branch — atau split jadi beberapa method khusus per filter kombo. Index dari Q10 akan partial-mitigate ORDER BY-nya."),

    ("Q12", "backtestRunRepository - findResearchLog", "backtest_run",
     "SELECT * FROM backtest_run WHERE status='COMPLETED' AND analysis_snapshot IS NOT NULL AND analysis_snapshot<>'' AND (user_id IS NULL OR user_id=?) ORDER BY created_time DESC",
     "CREATE INDEX idx_backtest_run_completed_with_snapshot ON backtest_run (created_time DESC) WHERE status='COMPLETED' AND analysis_snapshot IS NOT NULL;",
     "MEDIUM",
     "Dashboard /research. Sangat partial-friendly (status='COMPLETED' + snapshot not null). Partial index hemat storage + ordered scan + LIMIT pushdown. Akan jauh lebih kuat dari idx_backtest_run_status biasa pada query ini."),

    ("Q13", "backtestRunRepository - findByIdAndUserId", "backtest_run",
     "SELECT * FROM backtest_run WHERE backtest_run_id=? AND (user_id=? OR triggered_by='RESEARCHER')",
     None, "OK",
     "PK lookup + filter post-lookup. Pakai backtest_run_pkey. 1ms."),

    ("Q14", "backtestRunRepository - findRecentCompletedByAccountStrategyIdValueScan", "backtest_run",
     "SELECT * FROM backtest_run WHERE (account_strategy_id=? OR strategy_account_strategy_ids::text ILIKE '%uuid%') AND status='COMPLETED' AND ... ORDER BY created_time DESC LIMIT 5",
     "CREATE INDEX idx_backtest_run_as_id ON backtest_run (account_strategy_id) WHERE account_strategy_id IS NOT NULL; -- partial; JSONB branch tetap full-scan",
     "HIGH",
     "FEEDBACK DBA: branch ILIKE substring pada `strategy_account_strategy_ids::text` adalah **anti-index** — selalu Seq Scan tabel apa pun. JavaDoc sudah mengakui ini. Saat ini Seq Scan 1058 rows = 0.8ms, tapi akan jadi bottleneck cepat. Real fix bukan index biasa: (1) skema — pisahkan tabel `backtest_run_account_strategy` (junction table) jadi proper FK, atau (2) buat **GIN index** pada JSONB key 'strategyCode' via expression — `CREATE INDEX ... ON backtest_run USING GIN (strategy_account_strategy_ids jsonb_path_ops)` lalu ubah query pakai `?| array[uuid_text]` atau `@>` operator. Partial index pada account_strategy_id mempercepat branch utama; JSONB branch tetap perlu rework."),

    ("Q15", "backtestRunRepository - markAsHoldoutRun (UPDATE)", "backtest_run",
     "UPDATE backtest_run SET is_holdout_run=TRUE, holdout_for_sweep_id=? WHERE backtest_run_id=? AND is_holdout_run=FALSE",
     None, "OK",
     "Index Scan pada backtest_run_pkey. 0.4ms. Idx_backtest_run_holdout_per_sweep (UNIQUE WHERE is_holdout_run=true) sudah enforce constraint."),

    # ---- signal_history (33K rows hypertable, 4 indexes) ----
    ("Q16", "signalHistoryRepository - findLatest", "signal_history",
     "SELECT s FROM SignalHistory s WHERE signalId=? AND symbol=? ORDER BY ts DESC LIMIT 1",
     None, "OK",
     "Hypertable scan pakai compressed-segment metadata index. 0.6ms. PK (signal_id, symbol, ts) cover sempurna."),

    ("Q17", "signalHistoryRepository - findAt", "signal_history",
     "SELECT s FROM SignalHistory s WHERE signalId=? AND symbol=? AND ts=?",
     None, "OK",
     "Exact match pada PK. 4.3ms karena harus probe semua compressed chunks (TimescaleDB tidak bisa langsung chunk-prune pada exact ts saja). Untuk speed-up: pass-in range narrow di caller (ts BETWEEN ts-1 AND ts+1) supaya chunk pruning aktif. Bukan masalah index."),

    ("Q18", "signalHistoryRepository - findInRange", "signal_history",
     "SELECT s FROM SignalHistory s WHERE signalId=? AND symbol=? AND ts>=? AND ts<? ORDER BY ts ASC",
     None, "OK",
     "Range scan, chunk pruning aktif. 1.4ms untuk 1 bulan range. Optimal."),

    ("Q19", "signalHistoryRepository - countFor", "signal_history",
     "SELECT COUNT(*) FROM SignalHistory s WHERE signalId=? AND symbol=?",
     None, "OK",
     "Per-chunk count + sum. 0.6ms. Tidak butuh index baru."),

    # ---- inference_log (840 rows hypertable, 4 indexes) ----
    ("Q20", "inferenceLogRepository - findByIdAccountStrategyIdAndIdDecisionTimeBetween... (derived)", "inference_log",
     "SELECT * FROM inference_log WHERE account_strategy_id=? AND decision_time BETWEEN ? AND ? ORDER BY decision_time DESC",
     None, "OK",
     "Pakai idx_inference_log_strategy_time (account_strategy_id, decision_time DESC). Cover sempurna untuk filter + sort. 0.36ms."),

    ("Q21", "inferenceLogRepository - findReadyForBackfill", "inference_log",
     "SELECT * FROM inference_log WHERE realized_return IS NULL AND horizon_bars IS NOT NULL AND decision_time<=? ORDER BY decision_time ASC LIMIT ?",
     None, "OK",
     "Pakai partial idx_inference_log_unbackfilled (decision_time) WHERE realized_return IS NULL. Sempurna untuk backfill queue. 0.24ms."),

    ("Q22", "inferenceLogRepository - existsByIdDecisionTimeAndIdAccountStrategyId (derived)", "inference_log",
     "SELECT EXISTS(SELECT 1 FROM inference_log WHERE decision_time=? AND account_strategy_id=?)",
     None, "OK",
     "Index Only Scan pakai PK (decision_time, account_strategy_id). 0.01ms."),

    ("Q23", "inferenceLogRepository - findTradeReturnsForWindow", "inference_log",
     "SELECT realized_return FROM inference_log WHERE account_strategy_id=? AND decision_time>=? AND decision_time<? AND realized_return IS NOT NULL AND decision_type IN ('OPEN_LONG','OPEN_SHORT') ORDER BY decision_time ASC",
     None, "OK",
     "Pakai idx_inference_log_strategy_time. Filter decision_type + realized_return-NOT-NULL = bisa banyak Rows Removed by Filter saat data tumbuh, tapi cardinality OPEN_* trades tinggi. 0.02ms saat ini."),

    # ---- model_registry (19 rows, 5 indexes) ----
    ("Q24", "modelRegistryRepository - findLatestByStatus", "model_registry",
     "SELECT m FROM ModelRegistry m WHERE purpose=? AND symbol=? AND ((interval IS NULL AND m.interval IS NULL) OR m.interval=?) AND ((horizonBars IS NULL AND m.horizonBars IS NULL) OR m.horizonBars=?) AND m.status=? ORDER BY m.version DESC",
     None, "OK",
     "FEEDBACK DBA: query branch `IS NULL OR =` membuat indeks composite tidak terpakai sempurna. Saat ini Seq Scan 19 rows = 0.05ms. Index idx_model_registry_status_purpose (status, purpose, symbol) existing sudah cocok prefix sebagian, tapi Seq Scan tetap menang karena tabel tiny. Tidak perlu tuning kecuali tabel tumbuh > 1000 rows."),

    ("Q25", "modelRegistryRepository - findLatestByStatusSynced", "model_registry",
     "SELECT m FROM ModelRegistry m WHERE purpose=? AND symbol=? AND ... AND status=? AND artifact_synced_to_vps=true ORDER BY version DESC",
     None, "OK",
     "Hot-path live serving. Saat ini Seq Scan 19 rows. Bila tabel tumbuh: tambah partial index `(purpose, symbol, interval, horizon_bars) WHERE status='live' AND artifact_synced_to_vps=true` (mirror existing idx_model_registry_live tapi tambah filter sync). Tidak urgent."),

    ("Q26", "modelRegistryRepository - findByArtifactSha256 (derived)", "model_registry",
     "SELECT * FROM model_registry WHERE artifact_sha256=?",
     None, "OK",
     "UNIQUE INDEX uq_model_registry_artifact_sha256 existing. Saat ini Seq Scan 19 rows (cost menang), tapi akan otomatis switch ke index begitu volume bertambah."),

    ("Q27", "modelRegistryRepository - findByStatusOrderByCreatedTimeDesc (derived)", "model_registry",
     "SELECT * FROM model_registry WHERE status=? ORDER BY created_time DESC",
     None, "OK",
     "Seq Scan + Sort 19 rows. Index idx_model_registry_status_purpose existing serve filter status; sort by created_time tidak ter-cover tapi tabel kecil. OK."),

    # ---- funding_rate_history (14K rows, PK only) ----
    ("Q28", "fundingRateRepository - findLatestBefore", "funding_rate_history",
     "SELECT fr FROM FundingRate fr WHERE symbol=? AND fundingTime<=? ORDER BY fundingTime DESC",
     None, "OK",
     "PK (symbol, funding_time) sempurna untuk backward scan. 0.026ms. Catatan: default method `findLatest` di repo memanggil yang ini lalu ambil index[0] — kalau jumlah row besar boros memory; sebaiknya wrapper langsung pakai Pageable PageRequest.of(0,1) atau LIMIT 1 di query (kode side-fix, bukan DBA)."),

    ("Q29", "fundingRateRepository - findInWindow", "funding_rate_history",
     "SELECT fr FROM FundingRate fr WHERE symbol=? AND fundingTime>? AND fundingTime<=? ORDER BY fundingTime ASC",
     None, "OK",
     "PK range scan, 0.007ms. Cover sempurna."),

    ("Q30", "fundingRateRepository - findAllUpTo", "funding_rate_history",
     "SELECT fr FROM FundingRate fr WHERE symbol=? AND fundingTime<=? ORDER BY fundingTime ASC",
     None, "OK",
     "PK range scan ascending, 6986 rows = 0.9ms. Used by bulk backfill — tidak hot-path live. OK."),
]

assert len(QUERIES) == 30

# ----- Build Excel -----
wb = Workbook()
ws = wb.active
ws.title = "Tuning - trading_db (batch 2)"

HEADERS = [
    ("No", 5),
    ("Repository - Method", 55),
    ("Target Table(s)", 24),
    ("Query (ringkas)", 70),
    ("Current Plan (top op)", 42),
    ("Exec Time (ms)", 12),
    ("Issues", 30),
    ("Recommended Index DDL", 90),
    ("Priority", 10),
    ("Remarks", 55),
    ("Feedback DBA", 78),
]

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
priority_fills = {
    "HIGH":   PatternFill("solid", fgColor="F8CBAD"),
    "MEDIUM": PatternFill("solid", fgColor="FFE699"),
    "LOW":    PatternFill("solid", fgColor="E2EFDA"),
    "OK":     PatternFill("solid", fgColor="DDEBF7"),
}

for col_idx, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col_idx, value=label)
    c.fill, c.font = header_fill, header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "C2"

ISSUE_MESSAGES = {
    "external_merge": "Sort spill ke disk",
    "seq_scan_present": "Seq Scan terdeteksi",
    "bitmap_heap": "Bitmap Heap Scan",
    "rows_removed_high": "Banyak Rows Removed by Filter",
}

for i, (qid, repo_method, target_table, sql, ddl, prio, dba_note) in enumerate(QUERIES, start=2):
    p = parsed[qid]
    issues = [msg for flag, msg in ISSUE_MESSAGES.items() if p["flags"].get(flag)]
    # Suppress "Seq Scan" issue for tiny tables (legitimate optimizer choice)
    if target_table in {"account_strategy", "model_registry"} and "Seq Scan terdeteksi" in issues:
        issues = [i for i in issues if i != "Seq Scan terdeteksi"]
        issues.append("Seq Scan (table kecil — wajar)")
    issue_text = "; ".join(issues) if issues else "—"

    row_vals = [
        int(qid[1:]),
        repo_method,
        target_table,
        sql,
        p["top_op"][:200],
        round(p["exec_ms"], 3) if p["exec_ms"] is not None else None,
        issue_text,
        ddl if ddl else "(tidak ada perubahan)",
        prio,
        repo_method,
        dba_note,
    ]
    for col_idx, v in enumerate(row_vals, start=1):
        c = ws.cell(row=i, column=col_idx, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        c.border = border
        c.font = Font(size=10)
    ws.cell(row=i, column=9).fill = priority_fills.get(prio, PatternFill())
    ws.cell(row=i, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=9).font = Font(bold=True, size=10)
    ws.row_dimensions[i].height = 95

# ----- Ringkasan -----
ws2 = wb.create_sheet("Ringkasan")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 90

# Count priorities
prio_counts = {"HIGH": 0, "MEDIUM": 0, "OK": 0}
ddl_count = 0
for _, _, _, _, ddl, prio, _ in QUERIES:
    prio_counts[prio] = prio_counts.get(prio, 0) + 1
    if ddl:
        ddl_count += 1

summary_rows = [
    ("Database", "trading_db (PostgreSQL 16.13 + TimescaleDB 2.27)"),
    ("Batch", "2 — live decision path repositories"),
    ("Analisa tanggal", "2026-05-24"),
    ("Total query dianalisa", "30"),
    ("Repositori dicakup", "AccountStrategyRepository (8) · BacktestRunRepository (7) · SignalHistoryRepository (4) · InferenceLogRepository (4) · ModelRegistryRepository (4) · FundingRateRepository (3)"),
    ("", ""),
    ("Klasifikasi prioritas", ""),
    ("HIGH (wajib tindakan)", f"{prio_counts['HIGH']} — Q14 (JSONB substring scan)"),
    ("MEDIUM (rekomendasi)", f"{prio_counts['MEDIUM']} — Q02, Q10, Q11, Q12"),
    ("OK (no change)", f"{prio_counts['OK']}"),
    ("Index DDL yang ditawarkan", f"{ddl_count}"),
    ("", ""),
    ("Catatan kunci", ""),
    ("Q02 — temuan struktural", "JavaDoc menyebut 'uq_account_strategy' unique constraint, namun di pg_indexes tidak ditemukan. Tidak ada DB-level enforcement uniqueness pada (account_id, strategy_definition_id, symbol, interval_name). Rekomendasi: UNIQUE INDEX partial WHERE is_deleted=false."),
    ("Q14 — anti-index pattern", "JSONB substring ILIKE pada strategy_account_strategy_ids. Real fix bukan tambah btree biasa — perlu (a) skema relasional (junction table) atau (b) GIN index dengan operator @> / ?|."),
    ("Q24-Q27 model_registry", "Branch `IS NULL OR =` membuat composite index tidak optimal. Tapi tabel cuma 19 rows — Seq Scan menang. Tidak butuh tuning sampai volume tumbuh."),
    ("Q11 findFiltered", "15-param dynamic SQL dengan `CAST(:p AS TEXT) IS NULL OR ...`. Pattern sulit dioptimalkan via index. Mempertimbangkan Spring Data Specification untuk generate query bebas-NULL-branch (code change, bukan DBA)."),
    ("Q17 signal_history findAt", "Exact-ts match perlu probe semua chunks (chunk pruning gagal). 4.3ms. Caller bisa pass range narrow (ts ± 1 detik) untuk aktifkan chunk pruning — bukan masalah index."),
    ("Q28 fundingRate findLatest default", "Default wrapper Java memanggil List variant lalu ambil index[0]. Boros memory bila list besar. Sebaiknya tambah Pageable PageRequest.of(0,1) di caller (code change)."),
    ("", ""),
    ("Saran lifecycle berikutnya", ""),
    ("Setelah batch 2 di-review", "Aplikasi 3 index MEDIUM (idx_account_strategy_uniq_lookup, idx_backtest_run_created_time, idx_backtest_run_completed_with_snapshot) + diskusi rework Q14 (skema/GIN)."),
    ("Batch 3 (kalau lanjut)", "Backtest analysis path: BacktestTradeRepository, BacktestTradePositionRepository, BacktestEquityPointRepository, StrategyDailyRealizedCurveRepository, WalkForwardRunRepository, MonteCarloRunRepository."),
]
for r, (k, v) in enumerate(summary_rows, start=1):
    ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=11)
    ws2.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    if not k:
        ws2.cell(row=r, column=1).font = Font(size=11)
    if r <= 5 or k.startswith("HIGH") or k.startswith("MEDIUM") or k.startswith("OK"):
        ws2.row_dimensions[r].height = 22
    else:
        ws2.row_dimensions[r].height = 50

# ----- DDL sheet -----
ws3 = wb.create_sheet("Index DDL (siap-jalan)")
ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 120
ws3.cell(row=1, column=1, value="Q [Priority]").font = Font(bold=True)
ws3.cell(row=1, column=2, value="DDL — `CREATE INDEX CONCURRENTLY` aman ONLINE pada regular table").font = Font(bold=True)

ddls = []
for qid, repo_method, _, _, ddl, prio, _ in QUERIES:
    if ddl:
        # Convert CREATE INDEX -> CREATE INDEX CONCURRENTLY (regular tables; safe)
        ddl_concurrent = re.sub(r"CREATE (UNIQUE )?INDEX(?! CONCURRENTLY)", lambda m: f"CREATE {m.group(1) or ''}INDEX CONCURRENTLY", ddl, count=1)
        ddls.append((qid, repo_method, prio, ddl_concurrent))

for r, (qid, rm, prio, d) in enumerate(ddls, start=2):
    ws3.cell(row=r, column=1, value=f"{qid} [{prio}]").font = Font(size=10, bold=True)
    ws3.cell(row=r, column=2, value=f"-- {rm}\n{d}").alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    ws3.row_dimensions[r].height = 80

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"   Sheets: Tuning ({len(QUERIES)} rows) | Ringkasan | Index DDL ({len(ddls)} rows)")
print(f"   Priority breakdown: HIGH={prio_counts['HIGH']}, MEDIUM={prio_counts['MEDIUM']}, OK={prio_counts['OK']}")
