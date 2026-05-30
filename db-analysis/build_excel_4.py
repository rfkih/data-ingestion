"""Build batch-4 30-query DBA tuning Excel: Python services."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = Path(r"C:\Project\db-analysis\explain_results_4.txt")
OUT = Path(r"C:\Project\db-analysis\trading_db_tuning_local_30_batch4.xlsx")


def parse_explain_blocks(text):
    blocks = {}
    parts = re.split(r"^==== (Q\d+)\s+(.+)$", text, flags=re.MULTILINE)
    for i in range(1, len(parts), 3):
        qid, label, body = parts[i], parts[i + 1], parts[i + 2]
        m_exec = re.search(r"Execution Time:\s+([\d.]+)\s*ms", body)
        exec_ms = float(m_exec.group(1)) if m_exec else None
        plan_lines = []
        in_plan = False
        for ln in body.splitlines():
            if re.match(r"^-{5,}", ln):
                in_plan = True
                continue
            if in_plan:
                if ln.strip().startswith("Planning") or ln.strip().startswith("Execution Time") or "ERROR:" in ln:
                    break
                if ln.strip():
                    plan_lines.append(ln)
        top_op = plan_lines[0].strip() if plan_lines else ""
        top_op_clean = re.sub(r"\s*\(cost=.+", "", top_op)
        if "ERROR:" in body:
            top_op_clean = "(INSERT — EXPLAIN ANALYZE blocked by FK on test data; logical plan = Insert + index lookup)"
        flags = {
            "seq_scan_present": bool(re.search(r"->\s+(Parallel\s+)?Seq Scan on (?!compress_)", body))
                                or bool(re.match(r"^\s*Seq Scan on", plan_lines[0] if plan_lines else "")),
            "external_merge": bool(re.search(r"Sort Method:\s+external merge", body)),
            "bitmap_heap": bool(re.search(r"->\s+(Parallel\s+)?Bitmap Heap Scan", body)),
            "rows_removed_high": bool(re.search(r"Rows Removed by Filter:\s+(?:[1-9]\d{2,})", body)),
            "sort_present": bool(re.search(r"->\s+Sort\s+\(", body)),
        }
        blocks[qid] = {"label": label.strip(), "exec_ms": exec_ms, "top_op": top_op_clean, "flags": flags}
    return blocks


text = RESULTS.read_text(encoding="utf-8-sig", errors="ignore")
parsed = parse_explain_blocks(text)
assert len(parsed) == 30, f"Expected 30, got {len(parsed)}"

QUERIES = [
    # ── orchestrator/agent_state.py ──
    ("Q01", "orchestrator.repo.agent_state - _queue_counts", "research_queue",
     "SELECT status, COUNT(*) FROM research_queue GROUP BY status",
     None, "OK",
     "Seq Scan 58 rows + HashAggregate. 4.6ms — sebagian besar buffer cache warm-up. Tabel kecil, tidak perlu index baru."),

    ("Q02", "orchestrator.repo.agent_state - _last_iterations", "research_iteration_log",
     "SELECT iteration_id, ... FROM research_iteration_log ORDER BY created_time DESC LIMIT n",
     None, "OK",
     "Pakai idx_research_iteration_created_time. Incremental Sort 3.7ms — 6 rows total. Hot path session boot."),

    ("Q03", "orchestrator.repo.agent_state - _recent_sig_edge_ids", "research_iteration_log",
     "SELECT iteration_id FROM research_iteration_log WHERE statistical_verdict='SIGNIFICANT_EDGE' AND created_time>NOW()-INTERVAL ... ORDER BY created_time DESC",
     None, "OK",
     "Pakai idx_research_iteration_stat_verdict. 0.67ms. Sempurna."),

    ("Q04", "orchestrator.repo.agent_state - _active_hypotheses", "research_journal",
     "SELECT journal_id, ... FROM research_journal WHERE entry_type='HYPOTHESIS' AND status='ACTIVE' ORDER BY created_time DESC",
     "CREATE INDEX idx_research_journal_active_hypotheses ON research_journal (created_time DESC) WHERE entry_type='HYPOTHESIS' AND status='ACTIVE';",
     "MEDIUM",
     "FEEDBACK DBA: 17.6ms — query paling lambat di agent state boot. Plan: Bitmap Heap Scan via idx_research_journal_status_type + Sort. Sort dipakai karena composite (status, entry_type) tidak include created_time. Partial index pada (created_time DESC) WHERE entry_type='HYPOTHESIS' AND status='ACTIVE' akan eliminate Sort step + ukurannya kecil (HYPOTHESIS rows minoritas)."),

    ("Q05", "orchestrator.repo.agent_state - _last_run_summary", "research_journal",
     "SELECT ... FROM research_journal WHERE entry_type='RUN_SUMMARY' ORDER BY created_time DESC LIMIT 1",
     None, "OK",
     "Index Scan idx_research_journal_created_time. 0.79ms. Backward scan ambil paling baru."),

    ("Q06", "orchestrator.repo.agent_state - _pending_specialist_reviews", "research_journal",
     "SELECT ... FROM research_journal WHERE entry_type='IDEA_BACKLOG' AND status='ACTIVE' AND structured_data->>'kind'='specialist_review_request' ORDER BY created_time ASC",
     None, "OK",
     "Bitmap Heap via idx_research_journal_status_type + JSONB ->> filter post-lookup. 4ms. Existing idx_research_journal_structured_data (GIN) bisa optional cover JSONB key tapi cost saat ini OK."),

    ("Q07", "orchestrator.repo.agent_state - _recent_specialist_verdicts", "research_journal",
     "SELECT ... FROM research_journal WHERE entry_type='STRATEGY_OUTCOME' AND status='ACTIVE' AND structured_data->>'kind'=... AND created_time>NOW()-72h ORDER BY created_time DESC",
     None, "OK",
     "Index Scan idx_research_journal_created_time. 3.1ms. Time-window filter dominate."),

    ("Q08", "orchestrator.repo.agent_state - _last_null_screen_per_surface", "research_journal",
     "SELECT DISTINCT ON (strategy_code, instrument, interval_name) ... FROM research_journal WHERE entry_type='NULL_SCREEN_RESULT' ORDER BY ...",
     None, "OK",
     "Index Scan idx_research_journal_status_type. DISTINCT ON cheap karena rows sedikit. 0.05ms."),

    # ── orchestrator/services/tick.py ──
    ("Q09", "orchestrator.services.tick - _resolve_account_strategy", "account_strategy",
     "SELECT account_strategy_id, allow_long, allow_short FROM account_strategy WHERE strategy_code=? AND account_id=? AND is_deleted=false LIMIT 1",
     None, "OK",
     "Seq Scan 53 rows. Hot path tiap tick, tapi tabel kecil. Batch 2 sudah tambah idx_account_strategy_uniq_lookup yang sebagian cover, walaupun query ini tidak include strategy_definition_id. Saat tumbuh: index `(strategy_code, account_id) WHERE is_deleted=false` direkomendasi."),

    ("Q10", "orchestrator.services.tick - _poll_backtest_status", "backtest_run",
     "SELECT status FROM backtest_run WHERE backtest_run_id=?",
     None, "OK",
     "PK lookup. 0.3ms. Sempurna."),

    # ── orchestrator/repo/signals.py ──
    ("Q11", "orchestrator.repo.signals - list_signals", "signal_definition + model_registry + account_strategy",
     "SELECT sd.*, ... LEFT JOIN model_registry + correlated subquery array_agg(strategy_code) FROM account_strategy",
     None, "OK",
     "Nested Loop Left Join, semua tabel kecil. 0.1ms. Correlated subquery untuk bound_strategy_codes — efisien karena per-row hit cuma 53 acs rows."),

    ("Q12", "orchestrator.repo.signals - get_signal_health_counters", "signal_history",
     "SELECT MAX(ts), COUNT(*) FILTER (... 24h), COUNT(*) FILTER (... 7d) FROM signal_history WHERE signal_id=? AND ts>NOW()-30d",
     None, "OK",
     "Multi-filter aggregate via hypertable chunk pruning. 1.5ms. Bagus."),

    ("Q13", "orchestrator.repo.signals - list_firings", "signal_history",
     "SELECT * FROM signal_history WHERE signal_id=? AND ts>=? AND ts<? ORDER BY ts DESC",
     None, "OK",
     "Index Scan + Sort backward. 0.4ms. Pageable cover via LIMIT/OFFSET."),

    ("Q14", "orchestrator.repo.signals - daily_counts", "signal_history",
     "SELECT (ts AT TIME ZONE 'UTC')::date, COUNT(*) FROM signal_history WHERE signal_id=? AND ts>NOW()-30d GROUP BY 1 ORDER BY 1",
     None, "OK",
     "Finalize HashAggregate parallel. 0.78ms. GROUP BY date dengan time-window predicate — optimal."),

    # ── orchestrator/repo/models.py ──
    ("Q15", "orchestrator.repo.models - find_by_artifact_sha", "model_registry",
     "SELECT * FROM model_registry WHERE artifact_sha256=?",
     None, "OK",
     "Seq Scan 19 rows. uq_model_registry_artifact_sha256 ready, akan dipakai saat tumbuh. 0.24ms."),

    ("Q16", "orchestrator.repo.models - insert_model.next_version_subquery", "model_registry",
     "SELECT COALESCE(MAX(version)+1, 1) FROM model_registry WHERE purpose=? AND symbol=? AND interval IS NOT DISTINCT FROM ? AND horizon_bars IS NOT DISTINCT FROM ?",
     None, "OK",
     "FEEDBACK DBA: pattern `IS NOT DISTINCT FROM` membatasi optimizer untuk pakai index biasa, tapi tabel kecil (19 rows). Saat tumbuh > 1000 rows: existing uq_model_registry_directional (purpose, symbol, interval, horizon_bars, version) akan dipakai. 0.03ms."),

    ("Q17", "orchestrator.repo.models - list_models (synced=false)", "model_registry",
     "SELECT ... FROM model_registry WHERE TRUE AND status=? AND artifact_synced_to_vps=false ORDER BY created_time DESC",
     None, "OK",
     "Seq Scan 19 rows. Saat tumbuh: existing idx_model_registry_status_purpose serve filter. 0.03ms."),

    # ── orchestrator/repo/experiments.py ──
    ("Q18", "orchestrator.repo.experiments - find_existing_run", "experiment_run",
     "SELECT * FROM experiment_run WHERE spec_name=? AND spec_version=? AND git_sha=? AND dataset_sha=?",
     None, "OK",
     "Seq Scan 2.2ms. Existing idx_experiment_run_content_sha (presumably (git_sha, dataset_sha) or content-hash variant) sudah ada, tapi optimizer pilih Seq Scan karena tabel masih kecil. Begitu volume tumbuh, will switch."),

    ("Q19", "orchestrator.repo.experiments - update_summary_metric (UPDATE)", "experiment_run",
     "UPDATE experiment_run SET summary_metrics=summary_metrics||jsonb_build_object(?,?) WHERE run_id=?",
     None, "OK",
     "Seq Scan + UPDATE. PK index ready (experiment_run_pkey). 5ms — small table, Seq Scan masih menang. JSONB concat in-place via `||` operator."),

    # ── orchestrator/repo/hypothesis_audit.py ──
    ("Q20", "orchestrator.repo.hypothesis_audit - count_cumulative_trials", "hypothesis_audit",
     "SELECT COUNT(*) FROM hypothesis_audit WHERE strategy_code=? AND iteration_id IS NOT NULL",
     None, "OK",
     "Seq Scan 398 rows. 3.3ms. Existing idx_hypothesis_audit_strategy ada — wait, optimizer tidak pakainya. Karena filter strategy_code='DCB' sangat selektif tapi rows-removed-by-filter rendah. Bisa pakai partial index `(strategy_code) WHERE iteration_id IS NOT NULL` tapi marginal."),

    ("Q21", "orchestrator.repo.hypothesis_audit - count_data_universe_trials", "hypothesis_audit",
     "SELECT COUNT(*) FROM hypothesis_audit WHERE symbol=? AND interval_name=? AND iteration_id IS NOT NULL",
     None, "OK",
     "Pakai idx_hypothesis_audit_data_universe (symbol, interval_name). Bitmap Index Scan + Heap. 0.15ms. Sempurna."),

    ("Q22", "orchestrator.repo.hypothesis_audit - axis_has_discard", "hypothesis_audit",
     "SELECT ... FROM hypothesis_audit WHERE strategy_code=? AND axis_set_hash=? AND decision_verdict='DISCARD' ORDER BY created_time DESC LIMIT 1",
     None, "OK",
     "Pakai idx_hypothesis_audit_discarded (presumably partial WHERE decision_verdict='DISCARD'). 0.44ms."),

    # ── inference/services/streaming.py ──
    ("Q23", "inference.services.streaming - find_streaming_targets", "signal_definition + account_strategy",
     "SELECT DISTINCT ... FROM signal_definition JOIN account_strategy ON acs.ml_regime_signal_name=sd.name WHERE sd.status IN(...) AND acs.interval_name=ANY(?)",
     None, "OK",
     "Nested Loop + 2 Seq Scan tabel kecil. 0.05ms. Hot path tiap streaming tick — saat ini optimal."),

    ("Q24", "inference.services.streaming - find_gap_for_target", "signal_history + feature_values",
     "SELECT (SELECT MAX(ts) FROM signal_history WHERE signal_id=? AND symbol=?), (SELECT MAX(ts) FROM feature_values WHERE symbol=? AND interval=?)",
     None, "OK",
     "Dua sub-query MAX(ts) — hypertable chunk-aware backward scan. 0.4ms. Pakai compress_hyper indexes."),

    # ── inference/repo/models.py ──
    ("Q25", "inference.repo.models - fetch_signal_definition", "signal_definition",
     "SELECT ... FROM signal_definition WHERE signal_id=?",
     None, "OK",
     "PK lookup. < 0.1ms. Hot serving path."),

    ("Q26", "inference.repo.models - fetch_model_registry_row", "model_registry",
     "SELECT ... FROM model_registry WHERE id=?",
     None, "OK",
     "PK lookup. 0.23ms (Seq Scan saat ini karena 19 rows, akan otomatis switch ke PK)."),

    # ── ingest/features/persistence.py ──
    ("Q27", "ingest.features.persistence - start_run (feature_compute_run INSERT)", "feature_compute_run",
     "INSERT INTO feature_compute_run (...) VALUES (gen_random_uuid(), ...)",
     None, "OK",
     "Simple INSERT + PK + FK check. Existing idx_feature_compute_run_feature + idx_feature_compute_run_status untuk dashboard query. INSERT path tidak butuh tambahan index."),

    ("Q28", "ingest.features.persistence - write_values (feature_values UPSERT)", "feature_values",
     "INSERT INTO feature_values (...) VALUES (...) ON CONFLICT (feature_name, version, symbol, interval, ts) DO UPDATE",
     None, "OK",
     "Hot write path. ON CONFLICT pakai PK feature_values_pkey (feature_name, version, symbol, interval, ts). Executemany batch saat compute selesai. Optimal."),

    ("Q29", "ingest.features.persistence - _count_persisted_for_run", "feature_values",
     "SELECT COUNT(*) FROM feature_values WHERE compute_run_id=?",
     None, "OK",
     "Pakai idx_feature_values_run (compute_run_id). Reconciliation post-write. < 1ms typical."),

    # ── train/loader.py ──
    ("Q30", "train.loader - _read_per_bar_feature (training data range)", "feature_values",
     "SELECT ts, value FROM feature_values WHERE feature_name=? AND version=? AND symbol=? AND interval=? AND ts>=? AND ts<? ORDER BY ts ASC",
     None, "OK",
     "Multi-chunk hypertable scan dengan chunk pruning. PK leading (feature_name, version, symbol, interval, ts) menutup filter + sort sempurna. Hot path saat training data load — 1.4ms untuk 17-bulan range."),
]

assert len(QUERIES) == 30

# ----- Build Excel -----
wb = Workbook()
ws = wb.active
ws.title = "Tuning - trading_db (batch 4)"

HEADERS = [("No", 5), ("Module - Function", 62), ("Target Table(s)", 30),
           ("Query (ringkas)", 75), ("Current Plan (top op)", 42),
           ("Exec Time (ms)", 12), ("Issues", 32), ("Recommended Index DDL", 90),
           ("Priority", 10), ("Remarks", 62), ("Feedback DBA", 80)]

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
priority_fills = {"HIGH": PatternFill("solid", fgColor="F8CBAD"),
                  "MEDIUM": PatternFill("solid", fgColor="FFE699"),
                  "LOW": PatternFill("solid", fgColor="E2EFDA"),
                  "OK": PatternFill("solid", fgColor="DDEBF7")}

for col_idx, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col_idx, value=label)
    c.fill, c.font = header_fill, header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "C2"

ISSUE_MESSAGES = {
    "external_merge": "Sort spill ke disk",
    "seq_scan_present": "Seq Scan terdeteksi",
    "bitmap_heap": "Bitmap Heap Scan",
    "rows_removed_high": "Banyak Rows Removed by Filter",
    "sort_present": "Sort step (composite bisa eliminate)",
}

# small tables on Python-side (those where Seq Scan is fine)
small_tables = {"research_queue", "account_strategy", "model_registry",
                "experiment_run", "signal_definition", "model_registry + ..."}

for i, (qid, module_func, target_table, sql, ddl, prio, dba_note) in enumerate(QUERIES, start=2):
    p = parsed[qid]
    issues = [msg for flag, msg in ISSUE_MESSAGES.items() if p["flags"].get(flag)]
    if any(t in target_table for t in small_tables) and "Seq Scan terdeteksi" in issues:
        issues = [x for x in issues if x != "Seq Scan terdeteksi"]
        issues.append("Seq Scan (table kecil — wajar)")
    issue_text = "; ".join(issues) if issues else "—"
    row_vals = [int(qid[1:]), module_func, target_table, sql, p["top_op"][:200],
                round(p["exec_ms"], 3) if p["exec_ms"] is not None else None,
                issue_text, ddl if ddl else "(tidak ada perubahan)", prio, module_func, dba_note]
    for col_idx, v in enumerate(row_vals, start=1):
        c = ws.cell(row=i, column=col_idx, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        c.border = border
        c.font = Font(size=10)
    ws.cell(row=i, column=9).fill = priority_fills.get(prio, PatternFill())
    ws.cell(row=i, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=9).font = Font(bold=True, size=10)
    ws.row_dimensions[i].height = 95

# ----- Ringkasan -----
ws2 = wb.create_sheet("Ringkasan")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 90

prio_counts = {"HIGH": 0, "MEDIUM": 0, "OK": 0}
ddl_count = 0
for _, _, _, _, ddl, prio, _ in QUERIES:
    prio_counts[prio] = prio_counts.get(prio, 0) + 1
    if ddl:
        ddl_count += 1

summary_rows = [
    ("Database", "trading_db (PostgreSQL 16.13 + TimescaleDB 2.27)"),
    ("Batch", "4 — Python services (research-orchestrator + ingest + inference + train)"),
    ("Analisa tanggal", "2026-05-24"),
    ("Total query dianalisa", "30"),
    ("Service dicakup", "research-orchestrator (22) · inference (4) · ingest (3) · train (1)"),
    ("File dicakup", "agent_state.py · tick.py · signals.py · models.py · experiments.py · hypothesis_audit.py · streaming.py · inference/repo/models.py · ingest/features/persistence.py · train/loader.py"),
    ("", ""),
    ("Klasifikasi prioritas", ""),
    ("HIGH (wajib tindakan)", f"{prio_counts['HIGH']} — (tidak ada — index coverage Python-side schema sudah excellent)"),
    ("MEDIUM (rekomendasi)", f"{prio_counts['MEDIUM']} — Q04 (research_journal active hypotheses)"),
    ("OK (no change)", f"{prio_counts['OK']}"),
    ("Index DDL yang ditawarkan", f"{ddl_count}"),
    ("", ""),
    ("Catatan kunci", ""),
    ("Python-side schema dirancang dengan baik", "research_journal punya 6 index (created_time, status_type, strategy_code, fulltext GIN, structured_data GIN). hypothesis_audit punya 5. experiment_run punya 6 (termasuk 3 GIN JSONB). research_iteration_log punya 6 termasuk uniqueness."),
    ("Q04 issue terbesar", "Active hypotheses query 17.6ms — Bitmap Heap + Sort. Partial index `(created_time DESC) WHERE entry_type='HYPOTHESIS' AND status='ACTIVE'` akan eliminate Sort step. Dipanggil tiap session boot agent."),
    ("Hypertable queries optimal", "signal_history queries (Q12-Q14, Q24): chunk pruning bekerja sempurna untuk time-window predicates. feature_values (Q28, Q29, Q30) pakai PK 5-kolom composite. Sub-millisecond untuk most cases."),
    ("UPSERT pattern bagus", "Q28 feature_values ON CONFLICT DO UPDATE pakai PK natural — tidak perlu separate idempotency check."),
    ("Tidak ada anti-pattern serius", "Tidak menemukan JSONB substring ILIKE seperti Q14 batch 2, dynamic OR-NULL ala Q11 batch 2, atau CASE-in-WHERE ala Q25 batch 3. Python services menulis SQL secara explicit dan targeted."),
    ("", ""),
    ("Total cumulative coverage", ""),
    ("Batch 1-4", "120 queries dianalisa total. 17 index applied (7+4+5+1?). Coverage Java trading-engine ~57%; Python services major paths covered."),
    ("Outstanding code-level items", "Q14 batch 2 (JSONB junction), Q25 batch 3 (severity_rank), Q30 batch 3 (Spring Specification), Q11 batch 2 (Spring Specification)."),
    ("", ""),
    ("Saran selanjutnya", ""),
    ("Wajib", "Apply 1 DDL Q04 ke local + prod (kecil tapi mempercepat session boot)."),
    ("Optional", "Audit final di repos sisa: StrategyDefinition (2), StrategyParam (3), AuditEvent, ResearchAgentActivity (2), TradeExecutionLog, ResearchControl, dst. Itu adalah ~30 query @Query yang belum disentuh."),
]
for r, (k, v) in enumerate(summary_rows, start=1):
    ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=11)
    ws2.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    if not k:
        ws2.cell(row=r, column=1).font = Font(size=11)
    if r <= 6 or k.startswith("HIGH") or k.startswith("MEDIUM") or k.startswith("OK") or k.startswith("Index"):
        ws2.row_dimensions[r].height = 22
    else:
        ws2.row_dimensions[r].height = 50

# ----- DDL sheet -----
ws3 = wb.create_sheet("Index DDL (siap-jalan)")
ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 120
ws3.cell(row=1, column=1, value="Q [Priority]").font = Font(bold=True)
ws3.cell(row=1, column=2, value="DDL — CREATE INDEX CONCURRENTLY (ONLINE, no lock)").font = Font(bold=True)

ddls = []
for qid, module_func, _, _, ddl, prio, _ in QUERIES:
    if ddl:
        ddl_c = re.sub(r"CREATE (UNIQUE )?INDEX(?! CONCURRENTLY)",
                       lambda m: f"CREATE {m.group(1) or ''}INDEX CONCURRENTLY", ddl, count=1)
        ddls.append((qid, module_func, prio, ddl_c))

for r, (qid, mf, prio, d) in enumerate(ddls, start=2):
    ws3.cell(row=r, column=1, value=f"{qid} [{prio}]").font = Font(size=10, bold=True)
    ws3.cell(row=r, column=2, value=f"-- {mf}\n{d}").alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    ws3.row_dimensions[r].height = 80

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"   Sheets: Tuning ({len(QUERIES)} rows) | Ringkasan | Index DDL ({len(ddls)} rows)")
print(f"   Priority: HIGH={prio_counts['HIGH']}, MEDIUM={prio_counts['MEDIUM']}, OK={prio_counts['OK']}")
