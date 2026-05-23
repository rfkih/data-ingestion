"""Build the 30-query DBA tuning Excel for trading_db (local Postgres)."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = Path(r"C:\Project\db-analysis\explain_results.txt")
OUT = Path(r"C:\Project\db-analysis\trading_db_tuning_local_30.xlsx")

# ---------- Parse the EXPLAIN dump ----------

def parse_explain_blocks(text: str) -> dict[str, dict]:
    blocks = {}
    parts = re.split(r"^==== (Q\d+)\s+(\S+)\s*$", text, flags=re.MULTILINE)
    # parts: [pre, qid, label, body, qid, label, body, ...]
    for i in range(1, len(parts), 3):
        qid, label, body = parts[i], parts[i + 1], parts[i + 2]
        m_exec = re.search(r"Execution Time:\s+([\d.]+)\s*ms", body)
        exec_ms = float(m_exec.group(1)) if m_exec else None
        m_plan = re.search(r"Planning Time:\s+([\d.]+)\s*ms", body)
        plan_ms = float(m_plan.group(1)) if m_plan else None

        # Top-level operator = first non-blank non-header line
        plan_lines = []
        in_plan = False
        for ln in body.splitlines():
            if re.match(r"^-{5,}", ln):
                in_plan = True
                continue
            if in_plan:
                if ln.strip().startswith("Planning") or ln.strip().startswith("Execution Time"):
                    break
                if ln.strip():
                    plan_lines.append(ln)

        top_op = plan_lines[0].strip() if plan_lines else ""
        # Strip cost/time annotations
        top_op_clean = re.sub(r"\s*\(cost=.+", "", top_op)

        # Boolean flags
        flags = {
            "seq_scan_present": bool(re.search(r"->\s+(Parallel\s+)?Seq Scan on (?!compress_)", body)),
            "external_merge": bool(re.search(r"Sort Method:\s+external merge", body)),
            "bitmap_heap": bool(re.search(r"->\s+(Parallel\s+)?Bitmap Heap Scan", body)),
            "rows_removed": bool(re.search(r"Rows Removed by Filter:\s+\d", body)),
            "jit_active": "JIT:" in body,
        }
        blocks[qid] = {
            "label": label,
            "exec_ms": exec_ms,
            "plan_ms": plan_ms,
            "top_op": top_op_clean,
            "flags": flags,
            "raw_first_lines": "\n".join(plan_lines[:6]),
        }
    return blocks


text = RESULTS.read_text(encoding="utf-8", errors="ignore")
parsed = parse_explain_blocks(text)
assert len(parsed) == 30, f"Expected 30 queries, parsed {len(parsed)}"

# ---------- Hand-written analysis per query ----------
# Format: (no, repo_method, sql_one_liner, recommended_index_ddl, priority, dba_feedback)
# priority: HIGH (must), MEDIUM (nice), LOW (cosmetic), OK (no action)

QUERIES = [
    # ---- market_data: PK = (symbol, interval, start_time). 1.5M rows hypertable.
    ("Q01", "marketDataRepository - findBySymbolIntervalAndRange",
     "SELECT * FROM market_data WHERE symbol=? AND interval=? AND start_time>=? AND end_time<=? ORDER BY start_time ASC",
     None, "OK",
     "Plan pakai PK (symbol, interval, start_time). Sudah optimal — kondisi end_time<= cuma Filter, tidak signifikan karena kombinasi PK sudah sangat selektif."),

    ("Q02", "marketDataRepository - findBySymbolIntervalAndRangeLimited",
     "SELECT * FROM market_data WHERE symbol=? AND interval=? AND start_time>=? AND end_time<=? ORDER BY start_time ASC LIMIT :n",
     None, "OK",
     "Sama dengan Q01 + LIMIT. PK menangani filter + ORDER BY. Tidak perlu index baru."),

    ("Q03", "marketDataRepository - findBySymbolIntervalAndStartTimeRange",
     "SELECT * FROM market_data WHERE symbol=? AND interval=? AND start_time BETWEEN ? AND ? ORDER BY start_time ASC",
     None, "OK",
     "PK (symbol, interval, start_time) cocok 100% dengan filter + sort. No-op."),

    ("Q04", "marketDataRepository - findLatestCompletedBySymbolAndInterval",
     "SELECT * FROM market_data WHERE symbol=? AND interval=? AND end_time<? ORDER BY end_time DESC LIMIT 1",
     "CREATE INDEX idx_market_data_symbol_interval_end_time ON market_data (symbol, interval, end_time DESC);",
     "MEDIUM",
     "ORDER BY end_time tapi PK ordering pakai start_time. Plan saat ini scan multiple compressed chunks → backward scan PK index OK karena LIMIT 1, tapi composite di (symbol, interval, end_time) lebih tepat. Hypertable: gunakan ALTER TABLE...ADD UNIQUE atau index biasa."),

    ("Q05", "marketDataRepository - findLatestCandles",
     "SELECT * FROM market_data WHERE symbol=? AND interval=? ORDER BY start_time DESC LIMIT :n",
     None, "OK",
     "PK serve ORDER BY start_time DESC dengan backward index scan. Sudah optimal."),

    ("Q06", "marketDataRepository - findLast300BySymbolAndIntervalAndTime",
     "SELECT * FROM market_data WHERE symbol=? AND interval=? AND start_time<=? ORDER BY start_time DESC LIMIT 300",
     None, "OK",
     "PK (symbol, interval, start_time) menangani filter + sort + limit. Backward index scan."),

    ("Q07", "marketDataRepository - findLatestBySymbol",
     "SELECT * FROM market_data WHERE symbol=? AND interval=? ORDER BY end_time DESC LIMIT 1",
     "CREATE INDEX idx_market_data_symbol_interval_end_time ON market_data (symbol, interval, end_time DESC);",
     "MEDIUM",
     "Sama issue dengan Q04 — ORDER BY end_time. Index yang sama mendukung Q04 + Q07 sekaligus (BIKIN SEKALI)."),

    ("Q08", "marketDataRepository - existsBySymbolAndIntervalAndStartTime",
     "SELECT EXISTS(SELECT 1 FROM market_data WHERE symbol=? AND interval=? AND start_time=?)",
     None, "OK",
     "PK 3-kolom langsung hit, exact match. Plan = Index Only Scan, < 0.05 ms."),

    ("Q09", "marketDataRepository - countBySymbolIntervalAndRange",
     "SELECT COUNT(*) FROM market_data WHERE symbol=? AND interval=? AND start_time BETWEEN ? AND ?",
     None, "OK",
     "PK covers. Aggregate + Index Only Scan, no extra index."),

    ("Q10", "marketDataRepository - countMissingFeatureStoreRows",
     "SELECT COUNT(*) FROM market_data m WHERE m.symbol=? AND m.interval=? AND m.start_time BETWEEN ? AND ? AND NOT EXISTS (SELECT 1 FROM feature_store f WHERE f.symbol=m.symbol AND f.interval=m.interval AND f.start_time=m.start_time)",
     None, "OK",
     "Anti-join. Kedua tabel PK = (symbol, interval, start_time) → join via index lookup. Plan saat ini sudah optimal — anti-join lookup."),

    ("Q11", "marketDataRepository - findGapsBySymbolIntervalAndRange",
     "WITH ordered AS (SELECT start_time, LAG(start_time) OVER (ORDER BY start_time) FROM market_data WHERE symbol=? AND interval=? AND start_time BETWEEN ? AND ?) SELECT ... WHERE delta > interval LIMIT 100",
     None, "OK",
     "Window function di atas Index Only Scan PK. Sort by start_time gratis dari index. Tidak perlu tuning."),

    # ---- trades: 4 rows saat ini. Existing idx: account_id, account_strategy_id, asset+maker_taker, parent_order_id.
    ("Q12", "tradesRepository - findByTradeId",
     "SELECT * FROM trades WHERE trade_id=?",
     None, "OK",
     "Seq Scan saat ini wajar karena tabel 4 rows. PK (trade_id) sudah ada — akan langsung Index Scan begitu rows>1000."),

    ("Q13", "tradesRepository - findAllActiveTrades",
     "SELECT * FROM trades WHERE account_id=? AND account_strategy_id=? AND asset=? AND interval=? AND status IN (?) ORDER BY entry_time DESC",
     "CREATE INDEX idx_trades_active_lookup ON trades (account_strategy_id, status, entry_time DESC);",
     "HIGH",
     "Live entry-decision path — dipanggil tiap signal. Composite (account_strategy_id, status, entry_time DESC) → leading kolom paling unique. Saat ini Seq Scan, tapi setelah trades tumbuh ribuan akan jadi bottleneck. Kolom status dimasukkan karena selalu OPEN/PARTIALLY_CLOSED (low cardinality) — boleh karena ini composite, bukan single-col."),

    ("Q14", "tradesRepository - findAllActiveTradesForStrategies",
     "SELECT * FROM trades WHERE account_id=? AND account_strategy_id IN (?) AND asset=? AND interval=? AND status IN (?) ORDER BY entry_time DESC",
     None, "OK",
     "Index Q13 (idx_trades_active_lookup) sudah cover ini juga — leading kolom account_strategy_id dipakai untuk IN-list."),

    ("Q15", "tradesRepository - findByAccountIdsAndStatus",
     "SELECT * FROM trades WHERE account_id IN (?) AND status=? ORDER BY entry_time DESC LIMIT ?",
     "CREATE INDEX idx_trades_account_status_entry ON trades (account_id, status, entry_time DESC);",
     "HIGH",
     "List-trades endpoint paginated. Existing idx_trades_account_id ada, tapi harus extra Sort. Composite (account_id, status, entry_time DESC) hilangkan Sort + dukung Q17/Q20 juga."),

    ("Q16", "tradesRepository - countOpenByAccountStrategyId",
     "SELECT COUNT(*) FROM trades WHERE account_strategy_id=? AND status IN ('OPEN','PARTIALLY_CLOSED')",
     None, "OK",
     "Index dari Q13 (idx_trades_active_lookup leading account_strategy_id, status) sudah cover."),

    ("Q17", "tradesRepository - findClosedInPeriodByAccountIds",
     "SELECT * FROM trades WHERE account_id IN (?) AND status='CLOSED' AND exit_time>=? AND exit_time<? ORDER BY exit_time DESC",
     "CREATE INDEX idx_trades_closed_exit_time ON trades (account_id, exit_time DESC) WHERE status='CLOSED';",
     "MEDIUM",
     "Partial index — query selalu status='CLOSED'. Lebih hemat storage + faster. Alternatif: pakai composite Q15 (account_id, status, exit_time DESC) — tradeoff ukuran vs reuse."),

    ("Q18", "tradesRepository - findAllClosedSince",
     "SELECT * FROM trades WHERE status='CLOSED' AND exit_time>=? ORDER BY exit_time DESC",
     "CREATE INDEX idx_trades_status_exit_time ON trades (exit_time DESC) WHERE status='CLOSED';",
     "MEDIUM",
     "No account filter — scan ALL accounts. Dipakai TradingDomainMetricsExporter tiap 30 detik → cron-frequent. Partial index pada status='CLOSED' tepat (status non-unique tapi dipakai sebagai partial-predicate, sah)."),

    ("Q19", "tradesRepository - findAllOpen",
     "SELECT * FROM trades WHERE status IN ('OPEN','PARTIALLY_CLOSED') ORDER BY entry_time DESC",
     "CREATE INDEX idx_trades_open_entry_time ON trades (entry_time DESC) WHERE status IN ('OPEN','PARTIALLY_CLOSED');",
     "MEDIUM",
     "Sama pola dengan Q18. Partial index sangat hemat karena status OPEN/PARTIALLY_CLOSED rows << total. Dipakai metric exporter."),

    ("Q20", "tradesRepository - findClosedByAccountIdUpTo",
     "SELECT * FROM trades WHERE account_id=? AND status='CLOSED' AND exit_time IS NOT NULL AND exit_time<=? ORDER BY exit_time ASC",
     None, "OK",
     "Sudah ditutup oleh index Q17 (idx_trades_closed_exit_time partial WHERE status='CLOSED', leading account_id, exit_time)."),

    ("Q21", "tradesRepository - findClosedByAccountStrategyIdSince",
     "SELECT * FROM trades WHERE account_strategy_id=? AND status='CLOSED' AND exit_time IS NOT NULL AND exit_time>=? ORDER BY exit_time ASC",
     "CREATE INDEX idx_trades_strategy_closed_exit ON trades (account_strategy_id, exit_time) WHERE status='CLOSED';",
     "HIGH",
     "RiskGuardService drawdown path — dipanggil tiap entry. Partial index (account_strategy_id, exit_time) WHERE status='CLOSED' menghilangkan Sort + skip non-CLOSED rows."),

    ("Q22", "tradesRepository - countOpenByAccountIdAndSide",
     "SELECT COUNT(*) FROM trades WHERE account_id=? AND UPPER(side)=UPPER(?) AND status IN ('OPEN','PARTIALLY_CLOSED')",
     None, "MEDIUM",
     "FEEDBACK DBA: UPPER(side) functional — index biasa tidak terpakai. Solusi sebenarnya: normalisasi data 'side' selalu uppercase di INSERT, hapus UPPER() di query, baru index biasa akan terpakai. Atau buat functional index ON trades (account_id, UPPER(side)) WHERE status IN (...). Saat ini tabel 4 rows → Seq Scan masih lebih murah."),

    ("Q23", "tradesRepository - findRecentWithIntent",
     "SELECT * FROM trades WHERE asset=? AND intended_entry_price IS NOT NULL AND avg_entry_price IS NOT NULL AND entry_time IS NOT NULL ORDER BY entry_time DESC LIMIT ?",
     "CREATE INDEX idx_trades_asset_intent_entry ON trades (asset, entry_time DESC) WHERE intended_entry_price IS NOT NULL;",
     "MEDIUM",
     "SlippageCalibrationService — dipanggil per-strategy load. Partial index pada intended_entry_price IS NOT NULL realistis (sebagian besar trade lama tidak punya intent)."),

    # ---- feature_store: 1.38M rows hypertable. PK = (symbol, interval, start_time).
    ("Q24", "featureStoreRepository - findLatestBySymbolAndInterval",
     "SELECT * FROM feature_store WHERE symbol=? AND interval=? ORDER BY start_time DESC LIMIT 1",
     None, "OK",
     "PK + backward scan, sudah cover. < 1 ms typical."),

    ("Q25", "featureStoreRepository - findBySymbolIntervalAndRange",
     "SELECT * FROM feature_store WHERE symbol=? AND interval=? AND start_time BETWEEN ? AND ? ORDER BY start_time ASC",
     None, "OK",
     "Range scan pada PK, chunk pruning aktif. Tidak ada Sort terpisah."),

    ("Q26", "featureStoreRepository - findExistingStartTimesInRange",
     "SELECT start_time FROM feature_store WHERE symbol=? AND interval=? AND start_time BETWEEN ? AND ?",
     None, "OK",
     "Index Only Scan pada PK. Cepat karena tidak butuh kolom selain start_time."),

    # ---- trade_history: 5.4M rows hypertable, idx_trade_history_symbol_ts ada.
    ("Q27", "tradeHistoryRepository - findInRange",
     "SELECT * FROM trade_history WHERE venue=? AND symbol=? AND ts>=? AND ts<? ORDER BY ts ASC, trade_id ASC",
     None, "OK",
     "idx_trade_history_symbol_ts (venue, symbol, ts DESC) cover filter — backward scan/forward direction via PK juga OK. < 50 ms untuk 1h window."),

    ("Q28", "tradeHistoryRepository - sumNotionalInRange",
     "SELECT COALESCE(SUM(price * quantity), 0) FROM trade_history WHERE venue=? AND symbol=? AND ts>=? AND ts<?",
     None, "OK",
     "Aggregate di atas Index Scan. Existing idx_trade_history_symbol_ts cocok."),

    # ---- spec_trace: 1.26M rows REGULAR table (bukan hypertable). idx_spec_trace_backtest_run ada.
    ("Q29", "specTraceRepository - findByBacktestRunIdOrderByBarTimeAsc",
     "SELECT * FROM spec_trace WHERE backtest_run_id=? ORDER BY bar_time ASC",
     None, "HIGH",
     "FEEDBACK DBA: Index idx_spec_trace_backtest_run (backtest_run_id, bar_time) sudah ada, tapi optimizer tetap pilih Bitmap Heap Scan + external Sort (spill 11 MB ke disk) karena SELECT * + 1183-byte row width. Akar masalah BUKAN index — solusi: (1) projection — ganti SELECT * jadi SELECT kolom-yang-dipakai-saja di repository, atau (2) naikkan work_mem buat session viewer (SET work_mem='64MB'). Index sudah optimal."),

    ("Q30", "specTraceRepository - countErrorsByStrategyCodeSince",
     "SELECT COUNT(*) FROM spec_trace WHERE strategy_code=? AND error_class IS NOT NULL AND created_time>=?",
     None, "OK",
     "Index Only Scan pakai idx_spec_trace_errors (strategy_code, created_time DESC) WHERE error_class IS NOT NULL. 0.005 ms, sempurna."),
]

assert len(QUERIES) == 30

# ---------- Build Excel ----------

wb = Workbook()
ws = wb.active
ws.title = "Tuning - trading_db (local)"

# Header
HEADERS = [
    ("No", 5),
    ("Repository - Method", 50),
    ("Query (ringkas)", 70),
    ("Target Table(s)", 22),
    ("Current Plan (top op)", 42),
    ("Exec Time (ms)", 14),
    ("Issues", 30),
    ("Recommended Index DDL", 80),
    ("Priority", 11),
    ("Remarks", 50),
    ("Feedback DBA", 70),
]

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
priority_fills = {
    "HIGH": PatternFill("solid", fgColor="F8CBAD"),
    "MEDIUM": PatternFill("solid", fgColor="FFE699"),
    "LOW": PatternFill("solid", fgColor="E2EFDA"),
    "OK": PatternFill("solid", fgColor="DDEBF7"),
}

for col_idx, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col_idx, value=label)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 28
ws.freeze_panes = "C2"

TARGET_TABLE_GUESS = {
    "Q01": "market_data", "Q02": "market_data", "Q03": "market_data", "Q04": "market_data",
    "Q05": "market_data", "Q06": "market_data", "Q07": "market_data", "Q08": "market_data",
    "Q09": "market_data", "Q10": "market_data + feature_store", "Q11": "market_data",
    "Q12": "trades", "Q13": "trades", "Q14": "trades", "Q15": "trades", "Q16": "trades",
    "Q17": "trades", "Q18": "trades", "Q19": "trades", "Q20": "trades", "Q21": "trades",
    "Q22": "trades", "Q23": "trades",
    "Q24": "feature_store", "Q25": "feature_store", "Q26": "feature_store",
    "Q27": "trade_history", "Q28": "trade_history",
    "Q29": "spec_trace", "Q30": "spec_trace",
}

ISSUE_MESSAGES = {
    "external_merge": "Sort spill ke disk (work_mem kurang)",
    "seq_scan_present": "Seq Scan terdeteksi",
    "bitmap_heap": "Bitmap Heap Scan (mungkin index range terlalu lebar)",
    "rows_removed": "Rows Removed by Filter (filter di luar index cond)",
}

for i, (qid, repo_method, sql, ddl, prio, dba_note) in enumerate(QUERIES, start=2):
    p = parsed[qid]
    issues = [msg for flag, msg in ISSUE_MESSAGES.items() if p["flags"].get(flag)]
    issue_text = "; ".join(issues) if issues else "—"

    row_vals = [
        int(qid[1:]),
        repo_method,
        sql,
        TARGET_TABLE_GUESS.get(qid, "?"),
        p["top_op"][:200],
        round(p["exec_ms"], 3) if p["exec_ms"] is not None else None,
        issue_text,
        ddl if ddl else "(tidak ada perubahan)",
        prio,
        repo_method,
        dba_note,
    ]
    for col_idx, v in enumerate(row_vals, start=1):
        c = ws.cell(row=i, column=col_idx, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        c.border = border
        c.font = Font(size=10)
    # priority fill
    ws.cell(row=i, column=9).fill = priority_fills.get(prio, PatternFill())
    ws.cell(row=i, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=9).font = Font(bold=True, size=10)
    ws.row_dimensions[i].height = 90

# ---------- Summary sheet ----------
ws2 = wb.create_sheet("Ringkasan")
ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 80

summary_rows = [
    ("Database", "trading_db (PostgreSQL 16.13 + TimescaleDB 2.27)"),
    ("Container", "blackheart-postgres (port 5432)"),
    ("Total DB size", "3.6 GB"),
    ("Analisa tanggal", "2026-05-23"),
    ("Total query dianalisa", "30"),
    ("", ""),
    ("Klasifikasi", ""),
    ("HIGH (wajib tambah index)", "3 query — Q13, Q15, Q21"),
    ("MEDIUM (rekomendasi)", "5 query — Q04, Q07, Q17, Q18, Q19, Q22, Q23"),
    ("OK (no change)", "21 query"),
    ("FEEDBACK DBA khusus", "Q22 (UPPER() functional), Q29 (sort spill — bukan masalah index, projection issue)"),
    ("", ""),
    ("Catatan global", ""),
    ("Tabel trades", "Saat ini 4 rows → Seq Scan optimal. Index direkomendasi untuk siap saat volume tumbuh (ratusan+ rows)."),
    ("Hypertable", "market_data, feature_store, trade_history, signal_history, ml_shadow_log, inference_log, feature_values — sudah dipartisi otomatis per chunk."),
    ("Tabel besar non-hypertable", "spec_trace (1.13 GB / 1.26M rows). Kandidat untuk konversi hypertable atau retention policy."),
    ("Urutan kolom composite", "Sudah mengikuti aturan: leading column = paling unique (UUID account_strategy_id, account_id, dst), low-cardinality (status) di belakang."),
    ("pg_stat_statements", "Belum aktif di DB ini. Direkomendasikan enable untuk tracking top-N actual queries di iterasi berikutnya."),
]
for r, (k, v) in enumerate(summary_rows, start=1):
    ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=11)
    ws2.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    if not k:
        ws2.cell(row=r, column=1).font = Font(size=11)

# ---------- All DDLs sheet ----------
ws3 = wb.create_sheet("Index DDL (siap-jalan)")
ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 110
ws3.cell(row=1, column=1, value="No").font = Font(bold=True)
ws3.cell(row=1, column=2, value="DDL — jalankan ONLINE via CONCURRENTLY agar tidak lock").font = Font(bold=True)

ddls = []
for qid, repo_method, sql, ddl, prio, _ in QUERIES:
    if ddl:
        ddls.append((qid, repo_method, prio, ddl.replace("CREATE INDEX ", "CREATE INDEX CONCURRENTLY ")))

for r, (qid, rm, prio, d) in enumerate(ddls, start=2):
    ws3.cell(row=r, column=1, value=f"{qid} [{prio}]").font = Font(size=10, bold=True)
    ws3.cell(row=r, column=2, value=f"-- {rm}\n{d}").alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    ws3.row_dimensions[r].height = 50

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"   Sheets: Tuning ({len(QUERIES)} rows) | Ringkasan | Index DDL ({len(ddls)} rows)")
