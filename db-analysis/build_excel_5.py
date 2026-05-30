"""Build batch-5 30-query DBA tuning Excel: remaining trading-engine repos."""
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = Path(r"C:\Project\db-analysis\explain_results_5.txt")
OUT = Path(r"C:\Project\db-analysis\trading_db_tuning_local_30_batch5.xlsx")


def parse_explain_blocks(text):
    blocks = {}
    parts = re.split(r"^==== (Q\d+)\s+(.+)$", text, flags=re.MULTILINE)
    for i in range(1, len(parts), 3):
        qid, label, body = parts[i], parts[i + 1], parts[i + 2]
        m_exec = re.search(r"Execution Time:\s+([\d.]+)\s*ms", body)
        exec_ms = float(m_exec.group(1)) if m_exec else None
        plan_lines = []
        in_plan = False
        for ln in body.splitlines():
            if re.match(r"^-{5,}", ln):
                in_plan = True
                continue
            if in_plan:
                if ln.strip().startswith("Planning") or ln.strip().startswith("Execution Time"):
                    break
                if ln.strip():
                    plan_lines.append(ln)
        top_op = plan_lines[0].strip() if plan_lines else ""
        top_op_clean = re.sub(r"\s*\(cost=.+", "", top_op)
        flags = {
            "seq_scan_present": bool(re.search(r"->\s+(Parallel\s+)?Seq Scan on (?!compress_)", body))
                                or bool(re.match(r"^\s*Seq Scan on", plan_lines[0] if plan_lines else "")),
            "external_merge": bool(re.search(r"Sort Method:\s+external merge", body)),
            "bitmap_heap": bool(re.search(r"->\s+(Parallel\s+)?Bitmap Heap Scan", body)),
            "rows_removed_high": bool(re.search(r"Rows Removed by Filter:\s+(?:[1-9]\d{2,})", body)),
            "sort_present": bool(re.search(r"->\s+Sort\s+\(", body)),
        }
        blocks[qid] = {"label": label.strip(), "exec_ms": exec_ms, "top_op": top_op_clean, "flags": flags}
    return blocks


text = RESULTS.read_text(encoding="utf-8-sig", errors="ignore")
parsed = parse_explain_blocks(text)
assert len(parsed) == 30, f"Expected 30, got {len(parsed)}"

QUERIES = [
    # ── AccountRepository ──
    ("Q01", "accountRepository - findByIsActive", "accounts",
     "SELECT * FROM accounts WHERE is_active=?",
     None, "OK",
     "Seq Scan tabel 4 rows. is_active partial index belum ada — saat ini Seq Scan wajar. Saat tumbuh > 1000: partial `(account_id) WHERE is_active='1'` direkomendasi."),

    ("Q02", "accountRepository - findByAccountId", "accounts",
     "SELECT a.* FROM accounts a WHERE a.account_id=?",
     None, "OK",
     "PK Index Scan via accounts_pkey. 0.055ms."),

    ("Q03", "accountRepository - findByUserId", "accounts",
     "SELECT a.* FROM accounts a WHERE a.user_id=? ORDER BY a.created_time ASC",
     None, "OK",
     "Pakai uq_accounts_user_username_lower (sebagai covering index karena leading user_id). Atau bisa pakai idx_accounts_user_id. ORDER BY created_time eliminated karena 0 rows match — saat data ada akan butuh Sort. Tabel kecil, OK."),

    ("Q04", "accountRepository - existsByUserIdAndUsernameIgnoreCase", "accounts",
     "SELECT EXISTS(... WHERE user_id=? AND LOWER(username)=LOWER(?))",
     None, "OK",
     "Pakai functional index uq_accounts_user_username_lower (LOWER(username)). 0.01ms. Sempurna."),

    # ── UserRepository ──
    ("Q05", "userRepository - findByEmail", "users",
     "SELECT * FROM users WHERE email=? LIMIT 1",
     None, "OK",
     "Pakai idx_users_email. 0.022ms. Hot login path."),

    ("Q06", "userRepository - findByUserId", "users",
     "SELECT * FROM users WHERE user_id=? LIMIT 1",
     None, "OK",
     "Pakai users_pkey. 0.071ms."),

    ("Q07", "userRepository - updateLastLogin (UPDATE)", "users",
     "UPDATE users SET last_login_at=?, updated_time=NOW(), updated_by=? WHERE user_id=?",
     None, "OK",
     "Pakai users_pkey. 0.24ms. Tracking field update — sudah optimal."),

    # ── PortfolioRepository ──
    ("Q08", "portfolioRepository - findByAccountIdAndAsset", "portfolio",
     "SELECT * FROM portfolio WHERE account_id=? AND asset=? LIMIT 1",
     None, "OK",
     "Seq Scan tabel 4 rows. Existing index: portfolio_pkey saja. Saat tumbuh: composite `(account_id, asset)` recommended — query ini paling sering dijalankan."),

    ("Q09", "portfolioRepository - findAllByAccountIdAndAssetIn", "portfolio",
     "SELECT * FROM portfolio WHERE account_id=? AND asset IN(?) AND is_active='1'",
     None, "OK",
     "Seq Scan 4 rows. Saat tumbuh: idx pada (account_id) WHERE is_active='1' partial cukup."),

    ("Q10", "portfolioRepository - findAllByAccountId", "portfolio",
     "SELECT * FROM portfolio WHERE account_id=? AND is_active='1' ORDER BY asset ASC",
     "CREATE INDEX idx_portfolio_account_active ON portfolio (account_id, asset) WHERE is_active='1';",
     "MEDIUM",
     "Hot path: live portfolio rebalance. Saat tumbuh menjadi puluhan-ratusan-aset per account, partial index (account_id, asset) WHERE is_active='1' eliminate Sort + serve lookup. Saat ini Seq Scan tabel 4 rows OK."),

    ("Q11", "portfolioRepository - findAllByAccountIdIn", "portfolio",
     "SELECT * FROM portfolio WHERE account_id IN(?) AND is_active='1' ORDER BY asset ASC",
     None, "OK",
     "Index Q10 (idx_portfolio_account_active) cover ini juga via IN-list."),

    # ── HistoricalBackfillJobRepository ──
    ("Q12", "historicalBackfillJobRepository - findTopOrderByCreatedAtDesc", "historical_backfill_job",
     "SELECT * FROM historical_backfill_job ORDER BY created_at DESC LIMIT 20",
     None, "OK",
     "Pakai idx_historical_backfill_job_created. 0.032ms. Backward scan."),

    ("Q13", "historicalBackfillJobRepository - findTopByStatusOrderByCreatedAtDesc", "historical_backfill_job",
     "SELECT * FROM historical_backfill_job WHERE status=? ORDER BY created_at DESC LIMIT 20",
     None, "OK",
     "Pakai idx_historical_backfill_job_status. 0.01ms. Filter status selektif."),

    ("Q14", "historicalBackfillJobRepository - findActive", "historical_backfill_job",
     "SELECT * FROM historical_backfill_job WHERE status IN('PENDING','RUNNING') ORDER BY created_at DESC",
     None, "OK",
     "Seq Scan + Sort, 0.19ms / 3 rows. Saat job count tumbuh besar: partial `(created_at DESC) WHERE status IN('PENDING','RUNNING')` bisa eliminate Sort. Marginal."),

    # ── StrategyDefinitionRepository ──
    ("Q15", "strategyDefinitionRepository - findByStrategyCode (derived)", "strategy_definition",
     "SELECT * FROM strategy_definition WHERE strategy_code=?",
     None, "OK",
     "Pakai uq_strategy_definition_code. 0.044ms. UNIQUE lookup. Hot path."),

    ("Q16", "strategyDefinitionRepository - findFiltered (ILIKE search)", "strategy_definition",
     "SELECT * FROM strategy_definition WHERE :q IS NULL OR LOWER(strategy_code) LIKE ... OR LOWER(strategy_name) LIKE ... ORDER BY created_time DESC",
     None, "OK",
     "FEEDBACK DBA: pola `:p IS NULL OR ... LIKE '%pat%'` — anti-index pada substring (leading wildcard). Tapi tabel masih kecil. Saat tumbuh > 10K rows: pertimbangkan trigram index (pg_trgm extension + GIN). Tidak urgent."),

    # ── StrategyParamRepository ──
    ("Q17", "strategyParamRepository - findActiveByAccountStrategyId", "strategy_param",
     "SELECT * FROM strategy_param WHERE account_strategy_id=? AND is_active=true AND is_deleted=false",
     None, "OK",
     "Pakai idx_strategy_param_by_account_strategy. 0.031ms. Hot live path — preset resolve."),

    ("Q18", "strategyParamRepository - findLiveByAccountStrategyId", "strategy_param",
     "SELECT * FROM strategy_param WHERE account_strategy_id=? AND is_deleted=false ORDER BY is_active DESC, created_time ASC",
     None, "OK",
     "Index Scan Backward via idx_strategy_param_active_lookup + Incremental Sort. 0.5ms. UI listing."),

    ("Q19", "strategyParamRepository - deactivateOthers (UPDATE)", "strategy_param",
     "UPDATE strategy_param SET is_active=false WHERE account_strategy_id=? AND is_active=true AND param_id<>?",
     None, "OK",
     "Seq Scan + UPDATE. 3.8ms — terlihat banyak overhead trigger. Existing idx_strategy_param_active_lookup (account_strategy_id, is_active) ready saat tumbuh. Bulk-flip path tidak frequent."),

    # ── StrategyConfigRepository ──
    ("Q20", "strategyConfigRepository - findActiveBySymbol (JOIN)", "strategy_config + trend_following_config_detail",
     "SELECT sc.*, tfcd.* FROM strategy_config sc JOIN trend_following_config_detail tfcd ... WHERE strategy_name=? AND interval_name=? AND symbol=? AND status='ACTIVE' AND enabled=TRUE",
     "CREATE INDEX idx_strategy_config_lookup ON strategy_config (strategy_name, interval_name, symbol) WHERE status='ACTIVE' AND enabled=TRUE;",
     "MEDIUM",
     "Live strategy resolve path. Saat ini Seq Scan strategy_config + Index Scan on JOIN partner. Tabel kecil — saat tumbuh > 100 configs, partial index pada (strategy_name, interval_name, symbol) WHERE status='ACTIVE' AND enabled=TRUE eliminate Seq Scan."),

    ("Q21", "strategyConfigRepository - findActiveDefault (NULL symbol)", "strategy_config + trend_following_config_detail",
     "SELECT ... WHERE symbol IS NULL AND status='ACTIVE' AND enabled=TRUE ORDER BY version DESC LIMIT 1",
     None, "OK",
     "Pattern fallback bila simbol-spesifik tidak ada. Saat tumbuh: index Q20 dengan `WHERE status='ACTIVE' AND enabled=TRUE` cover branch ini juga (`symbol IS NULL` is a sub-case)."),

    # ── SymbolStrategyApprovalRepository ──
    ("Q22", "symbolStrategyApprovalRepository - findActivePair", "symbol_strategy_approval",
     "SELECT * FROM symbol_strategy_approval WHERE symbol=? AND strategy_code=? AND revoked_at IS NULL",
     None, "OK",
     "Pakai symbol_strategy_approval_symbol_idx. Existing symbol_strategy_approval_active_unique (partial WHERE revoked_at IS NULL) sangat tepat. 0.03ms."),

    ("Q23", "symbolStrategyApprovalRepository - findAllActive", "symbol_strategy_approval",
     "SELECT * FROM symbol_strategy_approval WHERE revoked_at IS NULL ORDER BY symbol ASC, strategy_code ASC",
     None, "OK",
     "Pakai symbol_strategy_approval_active_unique. 0.009ms. Hot read."),

    # ── StrategyPromotionLogRepository ──
    ("Q24", "strategyPromotionLogRepository - findFirstByAccountStrategyId...", "strategy_promotion_log",
     "SELECT * FROM strategy_promotion_log WHERE account_strategy_id=? ORDER BY created_time DESC LIMIT 1",
     None, "OK",
     "Pakai ix_promotion_log_account_strategy. 0.088ms. Sempurna."),

    ("Q25", "strategyPromotionLogRepository - findRecent", "strategy_promotion_log",
     "SELECT * FROM strategy_promotion_log ORDER BY created_time DESC LIMIT 50",
     None, "OK",
     "Seq Scan + Sort 19 rows. Tabel kecil. Existing idx pada created_time akan kick in saat tumbuh."),

    ("Q26", "strategyPromotionLogRepository - findRecentFiltered", "strategy_promotion_log",
     "SELECT * FROM strategy_promotion_log WHERE (?code IS NULL OR UPPER(strategy_code)=UPPER(?)) AND (?state IS NULL OR to_state=?) ORDER BY created_time DESC",
     None, "OK",
     "FEEDBACK DBA: pattern `(?p IS NULL OR col=?)` sama dengan Q11 batch 2 & Q30 batch 3 — anti-tuning. 19 rows current → Seq Scan OK. Sebaiknya pakai Specification builder saat volume bertambah."),

    # ── ResearchAgentActivityRepository ──
    ("Q27", "researchAgentActivityRepository - findBySessionIdOrderByCreatedTimeAsc", "research_agent_activity",
     "SELECT * FROM research_agent_activity WHERE session_id=? ORDER BY created_time ASC",
     None, "OK",
     "Index Scan Backward via idx_raa_created. 0.157ms. Existing idx_raa_session likely paired."),

    ("Q28", "researchAgentActivityRepository - findSessionSummaries", "research_agent_activity",
     "SELECT session_id, agent_name, MIN/MAX/COUNT/array_agg + multiple FILTER ... FROM research_agent_activity GROUP BY session_id, agent_name ORDER BY MAX(created_time) DESC LIMIT 20",
     None, "OK",
     "Heavy aggregate dengan multiple FILTER expression + JSONB ->> filter. Sort + Seq Scan + HashAggregate. 0.4ms / 136 rows. Saat tumbuh > 100K rows: pertimbangkan materialized view yang di-refresh per jam (operator-controlled). Bukan urgent — analytic surface, bukan hot path."),

    # ── MlIngestScheduleRepository ──
    ("Q29", "mlIngestScheduleRepository - findBySourceAndSymbol", "ml_ingest_schedule",
     "SELECT * FROM ml_ingest_schedule WHERE source=? AND ((?sym IS NULL AND symbol IS NULL) OR symbol=?)",
     None, "OK",
     "Pakai idx_ml_ingest_schedule_source. 0.35ms. Branch IS NOT DISTINCT FROM-style — tapi tabel kecil."),

    # ── LsrStrategyParamRepository ──
    ("Q30", "lsrStrategyParamRepository - findByAccountStrategyId", "lsr_strategy_param",
     "SELECT * FROM lsr_strategy_param WHERE account_strategy_id=?",
     None, "OK",
     "Pakai pk_lsr_strategy_param (account_strategy_id is PK). 0.009ms. Sempurna."),
]

assert len(QUERIES) == 30

# ----- Build Excel -----
wb = Workbook()
ws = wb.active
ws.title = "Tuning - trading_db (batch 5)"

HEADERS = [("No", 5), ("Repository - Method", 62), ("Target Table(s)", 38),
           ("Query (ringkas)", 75), ("Current Plan (top op)", 42),
           ("Exec Time (ms)", 12), ("Issues", 32), ("Recommended Index DDL", 90),
           ("Priority", 10), ("Remarks", 62), ("Feedback DBA", 80)]

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
priority_fills = {"HIGH": PatternFill("solid", fgColor="F8CBAD"),
                  "MEDIUM": PatternFill("solid", fgColor="FFE699"),
                  "LOW": PatternFill("solid", fgColor="E2EFDA"),
                  "OK": PatternFill("solid", fgColor="DDEBF7")}

for col_idx, (label, width) in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col_idx, value=label)
    c.fill, c.font = header_fill, header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "C2"

ISSUE_MESSAGES = {
    "external_merge": "Sort spill ke disk",
    "seq_scan_present": "Seq Scan terdeteksi",
    "bitmap_heap": "Bitmap Heap Scan",
    "rows_removed_high": "Banyak Rows Removed by Filter",
    "sort_present": "Sort step (composite bisa eliminate)",
}

small_tables = {"accounts", "users", "portfolio", "strategy_definition", "strategy_param",
                "strategy_config", "symbol_strategy_approval", "strategy_promotion_log",
                "research_agent_activity", "ml_ingest_schedule", "lsr_strategy_param",
                "historical_backfill_job"}

for i, (qid, repo_method, target_table, sql, ddl, prio, dba_note) in enumerate(QUERIES, start=2):
    p = parsed[qid]
    issues = [msg for flag, msg in ISSUE_MESSAGES.items() if p["flags"].get(flag)]
    if any(t in target_table for t in small_tables) and "Seq Scan terdeteksi" in issues:
        issues = [x for x in issues if x != "Seq Scan terdeteksi"]
        issues.append("Seq Scan (table kecil — wajar)")
    issue_text = "; ".join(issues) if issues else "—"
    row_vals = [int(qid[1:]), repo_method, target_table, sql, p["top_op"][:200],
                round(p["exec_ms"], 3) if p["exec_ms"] is not None else None,
                issue_text, ddl if ddl else "(tidak ada perubahan)", prio, repo_method, dba_note]
    for col_idx, v in enumerate(row_vals, start=1):
        c = ws.cell(row=i, column=col_idx, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        c.border = border
        c.font = Font(size=10)
    ws.cell(row=i, column=9).fill = priority_fills.get(prio, PatternFill())
    ws.cell(row=i, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=9).font = Font(bold=True, size=10)
    ws.row_dimensions[i].height = 95

# ----- Ringkasan -----
ws2 = wb.create_sheet("Ringkasan")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 90

prio_counts = {"HIGH": 0, "MEDIUM": 0, "OK": 0}
ddl_count = 0
for _, _, _, _, ddl, prio, _ in QUERIES:
    prio_counts[prio] = prio_counts.get(prio, 0) + 1
    if ddl:
        ddl_count += 1

summary_rows = [
    ("Database", "trading_db (PostgreSQL 16.13 + TimescaleDB 2.27)"),
    ("Batch", "5 — remaining trading-engine repositories (auth + strategy config + admin)"),
    ("Analisa tanggal", "2026-05-24"),
    ("Total query dianalisa", "30"),
    ("Repositori dicakup", "Account (4) · User (3) · Portfolio (4) · HistoricalBackfillJob (3) · StrategyDefinition (2) · StrategyParam (3) · StrategyConfig (2) · SymbolStrategyApproval (2) · StrategyPromotionLog (3) · ResearchAgentActivity (2) · MlIngestSchedule (1) · LsrStrategyParam (1)"),
    ("", ""),
    ("Klasifikasi prioritas", ""),
    ("HIGH (wajib tindakan)", f"{prio_counts['HIGH']} — (tidak ada)"),
    ("MEDIUM (rekomendasi)", f"{prio_counts['MEDIUM']} — Q10 (portfolio composite), Q20 (strategy_config partial)"),
    ("OK (no change)", f"{prio_counts['OK']}"),
    ("Index DDL yang ditawarkan", f"{ddl_count}"),
    ("", ""),
    ("Catatan kunci", ""),
    ("Mayoritas table sangat kecil", "accounts (4), users (1), portfolio (4), strategy_definition (~10), historical_backfill_job (~100), strategy_promotion_log (19), research_agent_activity (~136). Existing index sebagian besar sudah memenuhi pattern lookup."),
    ("Pola anti-index ditemukan lagi", "Q16 (LIKE leading-wildcard search), Q26 (IS NULL OR =). Sama persis pattern Q11 batch 2 dan Q30 batch 3. Solusi sama: trigram GIN extension untuk substring search, atau Spring Specification untuk dynamic OR-NULL."),
    ("UPDATE path semua optimal", "Q07 (users), Q19 (strategy_param deactivateOthers) — semua via PK / existing index. < 4ms even untuk bulk flip."),
    ("Authentication path well-indexed", "Q05 users.findByEmail pakai idx_users_email, Q06 users.findByUserId pakai users_pkey, Q04 accounts existsByUsername pakai functional index uq_accounts_user_username_lower. Login speed sub-millisecond."),
    ("", ""),
    ("Total cumulative final", ""),
    ("Batch 1-5", "150 queries dianalisa total. 18 index applied + 1 partial fix. Coverage trading-engine + Python services hampir lengkap untuk pola dominan."),
    ("Tail end (skip-able)", "Sisa repositori (BacktestRunStrategy, AuditEvent, ServerIpLog, EmailVerificationToken, PasswordResetToken, SupportMessage, dst.) total ~30 @Query — semua hot-cold rare-access dengan tabel kecil. Diminishing returns kalau diteruskan."),
    ("", ""),
    ("Saran selanjutnya", ""),
    ("Apply hari ini", "2 DDL (Q10, Q20) — portfolio + strategy_config partial index. Keduanya partial → ukuran kecil. Aman di-apply tanpa downtime."),
    ("Outstanding code/schema items", "Q14 batch 2 (JSONB junction table or GIN), Q25 batch 3 (severity_rank column migration), Q11+Q30 batch 2/3 + Q26 batch 5 (Spring Specification refactor), Q16 batch 5 (trigram extension untuk substring search)."),
]
for r, (k, v) in enumerate(summary_rows, start=1):
    ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=11)
    ws2.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    if not k:
        ws2.cell(row=r, column=1).font = Font(size=11)
    if r <= 6 or k.startswith("HIGH") or k.startswith("MEDIUM") or k.startswith("OK") or k.startswith("Index"):
        ws2.row_dimensions[r].height = 22
    else:
        ws2.row_dimensions[r].height = 55

# ----- DDL sheet -----
ws3 = wb.create_sheet("Index DDL (siap-jalan)")
ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 120
ws3.cell(row=1, column=1, value="Q [Priority]").font = Font(bold=True)
ws3.cell(row=1, column=2, value="DDL — CREATE INDEX CONCURRENTLY (ONLINE, no lock)").font = Font(bold=True)

ddls = []
for qid, repo_method, _, _, ddl, prio, _ in QUERIES:
    if ddl:
        ddl_c = re.sub(r"CREATE (UNIQUE )?INDEX(?! CONCURRENTLY)",
                       lambda m: f"CREATE {m.group(1) or ''}INDEX CONCURRENTLY", ddl, count=1)
        ddls.append((qid, repo_method, prio, ddl_c))

for r, (qid, rm, prio, d) in enumerate(ddls, start=2):
    ws3.cell(row=r, column=1, value=f"{qid} [{prio}]").font = Font(size=10, bold=True)
    ws3.cell(row=r, column=2, value=f"-- {rm}\n{d}").alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    ws3.row_dimensions[r].height = 80

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"   Sheets: Tuning ({len(QUERIES)} rows) | Ringkasan | Index DDL ({len(ddls)} rows)")
print(f"   Priority: HIGH={prio_counts['HIGH']}, MEDIUM={prio_counts['MEDIUM']}, OK={prio_counts['OK']}")
